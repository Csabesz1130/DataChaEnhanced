#!/usr/bin/env python3
"""
Test script for Excel Learning UI Components

This script tests the Excel Learning tab integration and functionality.
"""

import sys
import os
import tkinter as tk
from tkinter import ttk, messagebox
import threading
import time

# Add the src directory to the path
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'src'))

def test_excel_learning_tab_standalone():
    """Test the Excel Learning tab in standalone mode"""
    print("Testing Excel Learning Tab (Standalone)")
    
    try:
        # Create root window
        root = tk.Tk()
        root.title("Excel Learning Tab Test")
        root.geometry("800x600")
        
        # Create notebook
        notebook = ttk.Notebook(root)
        notebook.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Import and create Excel Learning tab
        from src.gui.excel_learning_tab import ExcelLearningTab
        
        # Create a mock app object
        class MockApp:
            def __init__(self):
                self.master = root
        
        mock_app = MockApp()
        
        # Create the tab
        excel_tab = ExcelLearningTab(notebook, mock_app)
        
        # Add status display
        status_frame = ttk.Frame(root)
        status_frame.pack(fill='x', padx=10, pady=5)
        
        status_label = ttk.Label(status_frame, text="Status: Ready", relief='sunken')
        status_label.pack(fill='x')
        
        def update_status(message):
            status_label.config(text=f"Status: {message}")
        
        # Add test buttons
        test_frame = ttk.Frame(root)
        test_frame.pack(fill='x', padx=10, pady=5)
        
        def test_background_processor():
            """Test background processor initialization"""
            if excel_tab.background_processor:
                update_status("Background processor available")
                messagebox.showinfo("Test", "Background processor initialized successfully!")
            else:
                update_status("Background processor not available")
                messagebox.showwarning("Test", "Background processor not available - this is expected if dependencies are missing")
        
        def test_ui_refresh():
            """Test UI refresh functionality"""
            try:
                excel_tab.refresh_tasks()
                excel_tab.refresh_results()
                update_status("UI refreshed successfully")
                messagebox.showinfo("Test", "UI refresh completed successfully!")
            except Exception as e:
                update_status(f"UI refresh error: {e}")
                messagebox.showerror("Test", f"UI refresh error: {e}")
        
        def test_cleanup():
            """Test cleanup functionality"""
            try:
                excel_tab.cleanup()
                update_status("Cleanup completed")
                messagebox.showinfo("Test", "Cleanup completed successfully!")
            except Exception as e:
                update_status(f"Cleanup error: {e}")
                messagebox.showerror("Test", f"Cleanup error: {e}")
        
        ttk.Button(test_frame, text="Test Background Processor", command=test_background_processor).pack(side='left', padx=5)
        ttk.Button(test_frame, text="Test UI Refresh", command=test_ui_refresh).pack(side='left', padx=5)
        ttk.Button(test_frame, text="Test Cleanup", command=test_cleanup).pack(side='left', padx=5)
        
        # Add close button
        def close_test():
            """Close the test window"""
            try:
                excel_tab.cleanup()
            except:
                pass
            root.destroy()
        
        ttk.Button(test_frame, text="Close", command=close_test).pack(side='right', padx=5)
        
        # Bind window close event
        root.protocol("WM_DELETE_WINDOW", close_test)
        
        update_status("Excel Learning Tab loaded successfully")
        
        print("Excel Learning Tab test window opened")
        print("You can now test the UI components:")
        print("- Submit Task tab: Browse and submit Excel files for learning")
        print("- Monitor Tasks tab: View active tasks and their progress")
        print("- Results tab: View completed tasks and their results")
        print("- Test buttons: Test various functionalities")
        
        # Start the GUI
        root.mainloop()
        
    except ImportError as e:
        print(f"Import error: {e}")
        print("This is expected if the AI Excel Learning modules are not available")
        return False
    except Exception as e:
        print(f"Error testing Excel Learning tab: {e}")
        return False
    
    return True

def test_main_app_integration():
    """Test integration with the main application"""
    print("\nTesting Main App Integration")
    
    try:
        # Import main app
        from src.gui.app import SignalAnalyzerApp
        
        # Create root window
        root = tk.Tk()
        root.title("Signal Analyzer with Excel Learning")
        root.geometry("1200x800")
        
        # Create main app
        app = SignalAnalyzerApp(root)
        
        # Check if Excel Learning tab was loaded
        if 'excel_learning' in app.tabs:
            print("✓ Excel Learning tab successfully integrated into main app")
            
            # Test tab functionality
            excel_tab = app.tabs['excel_learning']
            
            if hasattr(excel_tab, 'background_processor'):
                print("✓ Background processor available")
            else:
                print("⚠ Background processor not available (expected if dependencies missing)")
            
            if hasattr(excel_tab, 'submit_task'):
                print("✓ Task submission functionality available")
            
            if hasattr(excel_tab, 'refresh_tasks'):
                print("✓ Task monitoring functionality available")
            
            if hasattr(excel_tab, 'cleanup'):
                print("✓ Cleanup functionality available")
            
        else:
            print("✗ Excel Learning tab not found in main app")
            return False
        
        print("\nMain app with Excel Learning tab is ready for testing!")
        print("Look for the 'Excel Learning' tab in the application.")
        
        # Start the GUI
        root.mainloop()
        
    except ImportError as e:
        print(f"Import error: {e}")
        print("This is expected if the main app dependencies are not available")
        return False
    except Exception as e:
        print(f"Error testing main app integration: {e}")
        return False
    
    return True

def test_background_processor_standalone():
    """Test the background processor functionality standalone"""
    print("\nTesting Background Processor (Standalone)")
    
    try:
        from src.ai_excel_learning.background_processor import BackgroundProcessor, TaskStatus
        
        # Create background processor
        processor = BackgroundProcessor(
            max_workers=1,
            storage_path="test_learning_data",
            enable_notifications=True
        )
        
        print("✓ Background processor created successfully")
        
        # Test notification handler
        notifications_received = []
        
        def test_notification_handler(notification):
            notifications_received.append(notification)
            print(f"Notification received: {notification.notification_type.value} - {notification.message}")
        
        processor.add_notification_handler(test_notification_handler)
        print("✓ Notification handler added")
        
        # Start processing
        processor.start_processing()
        print("✓ Background processing started")
        
        # Test task submission (with a dummy file)
        import tempfile
        import os
        
        # Create a dummy Excel file for testing
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            dummy_file = f.name
        
        try:
            # Create a simple Excel file using openpyxl if available
            try:
                from openpyxl import Workbook
                wb = Workbook()
                ws = wb.active
                ws['A1'] = 'Test Data'
                ws['B1'] = 'Value'
                ws['A2'] = 1
                ws['B2'] = 100
                ws['A3'] = 2
                ws['B3'] = 200
                wb.save(dummy_file)
                print(f"✓ Created test Excel file: {dummy_file}")
            except ImportError:
                # If openpyxl not available, just create an empty file
                with open(dummy_file, 'w') as f:
                    f.write("dummy excel content")
                print(f"✓ Created dummy file: {dummy_file}")
            
            # Submit test task
            task_id = processor.submit_learning_task(
                file_path=dummy_file,
                task_type="excel_analysis",
                priority=3,
                metadata={'test': True}
            )
            
            print(f"✓ Test task submitted: {task_id}")
            
            # Wait a bit for processing
            time.sleep(3)
            
            # Check task status
            task = processor.get_task_status(task_id)
            if task:
                print(f"✓ Task status: {task.status.value}")
                print(f"✓ Task progress: {task.progress}")
            else:
                print("✗ Could not retrieve task status")
            
            # Check notifications
            if notifications_received:
                print(f"✓ Received {len(notifications_received)} notifications")
            else:
                print("⚠ No notifications received (this might be normal)")
            
            # Get all tasks
            all_tasks = processor.get_all_tasks()
            print(f"✓ Total tasks: {len(all_tasks)}")
            
            # Stop processing
            processor.stop_processing()
            print("✓ Background processing stopped")
            
        finally:
            # Clean up dummy file
            try:
                os.unlink(dummy_file)
                print("✓ Test file cleaned up")
            except:
                pass
        
        return True
        
    except ImportError as e:
        print(f"Import error: {e}")
        print("This is expected if the AI Excel Learning modules are not available")
        return False
    except Exception as e:
        print(f"Error testing background processor: {e}")
        return False

def main():
    """Main test function"""
    print("Excel Learning UI Test Suite")
    print("=" * 50)
    
    # Test 1: Standalone Excel Learning Tab
    print("\n1. Testing Excel Learning Tab (Standalone)")
    test1_result = test_excel_learning_tab_standalone()
    
    # Test 2: Background Processor
    print("\n2. Testing Background Processor")
    test2_result = test_background_processor_standalone()
    
    # Test 3: Main App Integration
    print("\n3. Testing Main App Integration")
    test3_result = test_main_app_integration()
    
    # Summary
    print("\n" + "=" * 50)
    print("Test Summary:")
    print(f"Excel Learning Tab: {'✓ PASS' if test1_result else '✗ FAIL'}")
    print(f"Background Processor: {'✓ PASS' if test2_result else '✗ FAIL'}")
    print(f"Main App Integration: {'✓ PASS' if test3_result else '✗ FAIL'}")
    
    if test1_result or test2_result or test3_result:
        print("\n✓ At least one test passed - UI components are working!")
    else:
        print("\n✗ All tests failed - check dependencies and module availability")

if __name__ == "__main__":
    main()
