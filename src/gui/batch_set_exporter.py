"""
Excel Exporter with Proper Number Formatting
Location: src/gui/batch_set_exporter.py

CHANGES MADE:
1. Added proper Excel number formatting
2. Ensured values are written as numbers, not strings
3. Set appropriate decimal places for different data types
4. Added Number format styling
"""

import os
import csv
import numpy as np
from pathlib import Path
import datetime
import re
from tkinter import filedialog, messagebox
import tkinter as tk
from tkinter import ttk
from openpyxl import Workbook
from openpyxl.styles import Font, NamedStyle
from openpyxl.utils import get_column_letter
from collections import defaultdict
from src.utils.logger import app_logger
from src.io_utils.io_utils import ATFHandler
from src.analysis.action_potential import ActionPotentialProcessor
from src.filtering.filtering import combined_filter

class BatchSetExporter:
    """Batch exporter with corrected integration calculation and proper Excel formatting"""
    
    def __init__(self, parent_app):
        self.parent_app = parent_app
        
        # Default integration ranges (same as GUI defaults)
        self.default_ranges = {
            'hyperpol_start': 0,
            'hyperpol_end': 200,
            'depol_start': 0,
            'depol_end': 200
        }
        
    def _setup_excel_styles(self, wb):
        """Setup number formatting styles for Excel"""
        try:
            # Create named styles for different number types
            
            # Style for integral values (2 decimal places)
            integral_style = NamedStyle(name="integral_number")
            integral_style.number_format = '0.00'
            
            # Style for current values (7 decimal places)
            current_style = NamedStyle(name="current_number")
            current_style.number_format = '0.0000000'
            
            # Style for time values (3 decimal places)
            time_style = NamedStyle(name="time_number")
            time_style.number_format = '0.000'
            
            # Style for voltage (integer)
            voltage_style = NamedStyle(name="voltage_number")
            voltage_style.number_format = '0'
            
            # Add styles to workbook (only if not already added)
            try:
                wb.add_named_style(integral_style)
                wb.add_named_style(current_style)
                wb.add_named_style(time_style)
                wb.add_named_style(voltage_style)
            except ValueError:
                # Styles already exist, that's fine
                pass
                
        except Exception as e:
            app_logger.warning(f"Could not setup Excel styles: {str(e)}")
        
    def export_folder_by_sets(self):
        """Export ATF files organized by sets to Excel files"""
        try:
            # Select folder
            folder_path = filedialog.askdirectory(
                title="Select folder containing ATF files"
            )
            
            if not folder_path:
                return False
            
            folder_path = Path(folder_path)
            
            # Find all ATF files
            atf_files = list(folder_path.glob("*.atf"))
            
            if not atf_files:
                messagebox.showwarning("No Files", "No ATF files found")
                return False
            
            # Organize by sets
            file_sets = self._organize_files_by_sets(atf_files)
            
            if not file_sets:
                messagebox.showerror("Error", "Could not parse file names")
                return False
            
            # Select output directory
            output_dir = filedialog.askdirectory(
                title="Select output directory"
            )
            
            if not output_dir:
                return False
            
            output_dir = Path(output_dir) / "ExportedSets"
            output_dir.mkdir(exist_ok=True)
            
            # Process each set
            results = []
            for set_number, files_info in file_sets.items():
                success = self._process_set(set_number, files_info, output_dir)
                if success:
                    results.append(f"✓ Set {set_number}: Success")
                else:
                    results.append(f"✗ Set {set_number}: Failed")
            
            # Show results
            messagebox.showinfo("Export Complete", "\n".join(results))
            return True
            
        except Exception as e:
            app_logger.error(f"Error in batch export: {str(e)}")
            messagebox.showerror("Error", f"Export failed: {str(e)}")
            return False
    
    def _organize_files_by_sets(self, atf_files):
        """Parse filenames and organize by set number"""
        file_sets = defaultdict(list)
        pattern = re.compile(r'(\d{8}_\d{3,4})_(\d+)_([+-]?\d+)mV\.atf$', re.IGNORECASE)

        def _fallback_parse(name: str):
            """Fallback parser if regex doesn't match"""
            try:
                base, ext = os.path.splitext(name)
                if ext.lower() != '.atf':
                    return None
                parts = base.split('_')
                if len(parts) < 4:
                    return None
                file_number = f"{parts[0]}_{parts[1]}"
                set_number = int(parts[2])
                voltage_str = parts[3]
                voltage = int(voltage_str.replace('mV', '').replace('mv', ''))
                return file_number, set_number, voltage
            except Exception:
                return None
        
        for file_path in atf_files:
            name = file_path.name
            match = pattern.match(name)
            if match:
                file_number = match.group(1)
                set_number = int(match.group(2))
                voltage = int(match.group(3))
            else:
                parsed = _fallback_parse(name)
                if parsed is None:
                    app_logger.warning(f"Could not parse filename: {name}")
                    continue
                file_number, set_number, voltage = parsed

            file_sets[set_number].append({
                'file_path': file_path,
                'filename': name,
                'file_number': file_number,
                'voltage': voltage,
                'sheet_name': file_number
            })
        
        # Sort files within sets
        for set_number in file_sets:
            file_sets[set_number].sort(key=lambda x: x['file_number'])
        
        return dict(file_sets)
    
    def _process_set(self, set_number, files_info, output_dir):
        """Process a single set and create Excel file"""
        try:
            # Create workbook
            wb = Workbook()
            
            # Setup Excel number formatting styles
            self._setup_excel_styles(wb)
            
            # Process each file
            processed_count = 0
            for file_info in files_info:
                if self._process_file_to_sheet(wb, file_info):
                    processed_count += 1
            
            if processed_count == 0:
                app_logger.error(f"No files processed for set {set_number}")
                return False
            
            # Remove default sheet
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])
            
            # Save workbook
            timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = output_dir / f"Set_{set_number}_{timestamp}.xlsx"
            wb.save(filename)
            
            app_logger.info(f"Set {set_number} exported: {processed_count} files")
            return True
            
        except Exception as e:
            app_logger.error(f"Error processing set {set_number}: {str(e)}")
            return False
    
    def _process_file_to_sheet(self, wb, file_info):
        """Process single file and add to workbook - with CORRECTED integration"""
        try:
            # Load file
            atf_handler = ATFHandler(file_info['file_path'])
            atf_handler.load_atf()
            
            time_data = atf_handler.get_column("Time")
            raw_data = atf_handler.get_column("#1")
            
            if time_data is None or raw_data is None:
                return False
            
            # Get sampling rate
            sampling_interval = atf_handler.get_sampling_rate()
            fs_hz = 1.0 / sampling_interval if sampling_interval > 0 else 10000.0
            
            # Apply filters (same as working version)
            filter_params = {
                'savgol_params': {'window_length': 101, 'polyorder': 3},
                'butter_params': {'cutoff': 2000, 'fs': fs_hz, 'order': 2}
            }
            filtered_data = combined_filter(raw_data, **filter_params)
            
            # Create processor
            processor_params = {
                'n_cycles': 2,
                't0': 20,
                't1': 100,
                't2': 100,
                't3': 1000,
                'V0': -80,
                'V1': -100,
                'V2': file_info['voltage'],
                'cell_area_cm2': 0.0001,
                'time_constant_ms': 10,
                'threshold_percent': 10,
                'use_alternative_method': True,
                'show_modified_peaks': True,
            }
            
            processor = ActionPotentialProcessor(
                data=filtered_data, 
                time_data=time_data, 
                params=processor_params
            )
            
            # Process data (call methods directly like your working version does)
            processor.baseline_correction_initial()
            processor.advanced_baseline_normalization()
            processor.process_orange_curve_with_spike_removal()
            processor.normalized_curve, processor.normalized_curve_times = processor.calculate_normalized_curve()
            processor.average_curve, processor.average_curve_times = processor.calculate_segment_average()
            
            # Generate purple curves
            result = processor.apply_average_to_peaks()
            if result and result[0] is not None:
                processor.modified_hyperpol = result[0]
                processor.modified_hyperpol_times = result[1]
                processor.modified_depol = result[2]
                processor.modified_depol_times = result[3]
            else:
                app_logger.error(f"Failed to generate purple curves for {file_info['filename']}")
                return False
            
            # FIXED: Calculate integrals with correct ranges and units
            hyperpol_integral, depol_integral = self._calculate_correct_integrals(processor)
            
            # Store integral values on processor (like your working version expects)
            total_integral = hyperpol_integral + depol_integral
            processor.integral_value = total_integral  # Store as number, not string
            processor.hyperpol_area = hyperpol_integral  # Store as number
            processor.depol_area = depol_integral  # Store as number
            processor.purple_integral_value = (
                f"Hyperpol: {hyperpol_integral:.2f} pC, "
                f"Depol: {depol_integral:.2f} pC"
            )
            
            app_logger.info(f"Calculated integrals for {file_info['filename']}: "
                          f"Total={total_integral:.2f}, Hyperpol={hyperpol_integral:.2f}, "
                          f"Depol={depol_integral:.2f}")
            
            # Add sheet with data
            self._write_sheet(wb, file_info, processor)
            return True
            
        except Exception as e:
            app_logger.error(f"Error processing {file_info['filename']}: {str(e)}")
            return False
    
    def _calculate_correct_integrals(self, processor):
        """
        Calculate integrals using the same method as the GUI - CORRECTED VERSION
        
        Returns:
            tuple: (hyperpol_integral, depol_integral) in pC
        """
        hyperpol_integral = 0.0
        depol_integral = 0.0
        
        try:
            # Check if we have the purple curves
            if (not hasattr(processor, 'modified_hyperpol') or 
                not hasattr(processor, 'modified_hyperpol_times') or
                processor.modified_hyperpol is None or
                processor.modified_hyperpol_times is None):
                app_logger.warning("No hyperpol purple curve available")
                return 0.0, 0.0
            
            if (not hasattr(processor, 'modified_depol') or 
                not hasattr(processor, 'modified_depol_times') or
                processor.modified_depol is None or
                processor.modified_depol_times is None):
                app_logger.warning("No depol purple curve available")
                return 0.0, 0.0
            
            # Get integration ranges (same as GUI defaults)
            hyperpol_start = self.default_ranges['hyperpol_start']  # 0
            hyperpol_end = self.default_ranges['hyperpol_end']      # 200
            depol_start = self.default_ranges['depol_start']        # 0
            depol_end = self.default_ranges['depol_end']            # 200
            
            # Validate ranges for hyperpol
            hyperpol_len = len(processor.modified_hyperpol)
            if hyperpol_end > hyperpol_len:
                hyperpol_end = hyperpol_len
                app_logger.warning(f"Hyperpol end range adjusted to {hyperpol_end}")
            
            # Validate ranges for depol
            depol_len = len(processor.modified_depol)
            if depol_end > depol_len:
                depol_end = depol_len
                app_logger.warning(f"Depol end range adjusted to {depol_end}")
            
            # Calculate hyperpol integral (same as GUI method)
            if hyperpol_start < hyperpol_len and hyperpol_end <= hyperpol_len:
                hyperpol_data = processor.modified_hyperpol[hyperpol_start:hyperpol_end]
                hyperpol_times = processor.modified_hyperpol_times[hyperpol_start:hyperpol_end]
                
                if len(hyperpol_data) > 1:
                    # Integration: pA * ms = pA·ms
                    # To convert to pC: pA·ms * 1e-12 A/pA * 1e-3 s/ms * 1e12 pC/C = pA·ms * 1e-3 = pA·ms / 1000
                    hyperpol_integral_raw = np.trapz(hyperpol_data, x=hyperpol_times * 1000)
                    hyperpol_integral = hyperpol_integral_raw / 1000.0  # Convert pA·ms to pC
                    
                    app_logger.debug(f"Hyperpol: range {hyperpol_start}-{hyperpol_end}, "
                                   f"raw integral: {hyperpol_integral_raw:.2f} pA·ms, "
                                   f"converted: {hyperpol_integral:.2f} pC")
            
            # Calculate depol integral (same as GUI method)
            if depol_start < depol_len and depol_end <= depol_len:
                depol_data = processor.modified_depol[depol_start:depol_end]
                depol_times = processor.modified_depol_times[depol_start:depol_end]
                
                if len(depol_data) > 1:
                    # Integration: pA * ms = pA·ms
                    # To convert to pC: pA·ms / 1000
                    depol_integral_raw = np.trapz(depol_data, x=depol_times * 1000)
                    depol_integral = depol_integral_raw / 1000.0  # Convert pA·ms to pC
                    
                    app_logger.debug(f"Depol: range {depol_start}-{depol_end}, "
                                   f"raw integral: {depol_integral_raw:.2f} pA·ms, "
                                   f"converted: {depol_integral:.2f} pC")
            
        except Exception as e:
            app_logger.error(f"Error calculating integrals: {str(e)}")
            return 0.0, 0.0
        
        return hyperpol_integral, depol_integral
    
    def _write_number_with_format(self, ws, cell_ref, value, format_type="general"):
        """Write a number to Excel with proper formatting"""
        try:
            # Convert to float if it's a string representation
            if isinstance(value, str):
                # Try to extract number from string like "3.56 pC"
                import re
                number_match = re.search(r'([+-]?\d+\.?\d*)', value)
                if number_match:
                    numeric_value = float(number_match.group(1))
                else:
                    # If no number found, write as text
                    ws[cell_ref] = value
                    return
            else:
                numeric_value = float(value)
            
            # Write the number
            ws[cell_ref] = numeric_value
            
            # Apply formatting based on type
            if format_type == "integral":
                ws[cell_ref].number_format = '0.00'
            elif format_type == "current":
                ws[cell_ref].number_format = '0.0000000'
            elif format_type == "time":
                ws[cell_ref].number_format = '0.000'
            elif format_type == "voltage":
                ws[cell_ref].number_format = '0'
            # else: leave as general format
            
        except (ValueError, TypeError) as e:
            # If conversion fails, write as text
            app_logger.warning(f"Could not convert {value} to number, writing as text")
            ws[cell_ref] = str(value)
    
    def _write_sheet(self, wb, file_info, processor):
        """Write data to Excel sheet with proper number formatting"""
        # Create sheet
        sheet_name = file_info['sheet_name']
        if sheet_name in wb.sheetnames:
            sheet_name = f"{sheet_name}_v2"
        ws = wb.create_sheet(title=sheet_name)
        
        row = 1
        
        # File info
        ws[f'A{row}'] = "Original ATF File:"
        ws[f'B{row}'] = file_info['filename']
        row += 1
        
        ws[f'A{row}'] = "V2 Voltage (mV):"
        self._write_number_with_format(ws, f'B{row}', file_info['voltage'], "voltage")
        row += 2
        
        # Overall integral - WRITE AS NUMBER
        ws[f'A{row}'] = "Overall Integral (pC):"
        self._write_number_with_format(ws, f'B{row}', processor.integral_value, "integral")
        ws[f'A{row}'].font = Font(bold=True)
        row += 2
        
        # Purple Hyperpol Section
        ws[f'A{row}'] = "PURPLE HYPERPOL CURVE"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = "Hyperpol Integral from the ap:"
        self._write_number_with_format(ws, f'B{row}', processor.hyperpol_area, "integral")
        row += 1
        
        # Headers
        ws[f'A{row}'] = "Index"
        ws[f'B{row}'] = "Hyperpol_pA"
        ws[f'C{row}'] = "Hyperpol_time_ms"
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].font = Font(bold=True)
        row += 1
        
        # Data - WRITE AS NUMBERS WITH PROPER FORMATTING
        if hasattr(processor, 'modified_hyperpol') and hasattr(processor, 'modified_hyperpol_times'):
            hyperpol = processor.modified_hyperpol
            hyperpol_times = processor.modified_hyperpol_times
            for i in range(len(hyperpol)):
                ws[f'A{row}'] = i + 1  # Index as integer
                
                # Write current value as number with 7 decimal places
                self._write_number_with_format(ws, f'B{row}', float(hyperpol[i]), "current")
                
                # Write time value as number with 3 decimal places
                self._write_number_with_format(ws, f'C{row}', float(hyperpol_times[i] * 1000), "time")
                
                row += 1
        row += 1
        
        # Purple Depol Section
        ws[f'A{row}'] = "PURPLE DEPOL CURVE"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        ws[f'A{row}'] = "Depol Integral from the ap:"
        self._write_number_with_format(ws, f'B{row}', processor.depol_area, "integral")
        row += 1
        
        # Headers
        ws[f'A{row}'] = "Index"
        ws[f'B{row}'] = "Depol_pA"
        ws[f'C{row}'] = "Depol_time_ms"
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].font = Font(bold=True)
        row += 1
        
        # Data - WRITE AS NUMBERS WITH PROPER FORMATTING
        if hasattr(processor, 'modified_depol') and hasattr(processor, 'modified_depol_times'):
            depol = processor.modified_depol
            depol_times = processor.modified_depol_times
            for i in range(len(depol)):
                ws[f'A{row}'] = i + 1  # Index as integer
                
                # Write current value as number with 7 decimal places
                self._write_number_with_format(ws, f'B{row}', float(depol[i]), "current")
                
                # Write time value as number with 3 decimal places
                self._write_number_with_format(ws, f'C{row}', float(depol_times[i] * 1000), "time")
                
                row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width


# Button class for integration
class BatchSetExportButton:
    """Batch set export button for integration with the main app"""
    
    def __init__(self, parent_app, toolbar_frame):
        self.parent_app = parent_app
        self.exporter = BatchSetExporter(parent_app)
        
        # Add export button to toolbar
        self.export_button = tk.Button(
            toolbar_frame, 
            text="📊 Export Sets", 
            command=self.export_sets,
            relief="raised",
            bg="#2ECC71",
            fg="white",
            font=("Arial", 9, "bold"),
            padx=8,
            pady=2
        )
        self.export_button.pack(side='left', padx=2)
        
        app_logger.info("Batch set export button added to toolbar")
    
    def export_sets(self):
        """Export ATF file sets"""
        try:
            self.exporter.export_folder_by_sets()
        except Exception as e:
            app_logger.error(f"Error in set export: {str(e)}")
            messagebox.showerror("Export Error", f"Failed to export sets:\n{str(e)}")


# This is the function being imported by your app
def add_set_export_to_toolbar(parent_app, toolbar_frame):
    """
    Add the batch set export button to an existing toolbar.
    This is the function your app is trying to import.
    
    Args:
        parent_app: The main application instance
        toolbar_frame: The tkinter frame where the button should be added
    
    Returns:
        BatchSetExportButton: The created button instance
    """
    return BatchSetExportButton(parent_app, toolbar_frame)

"""
Debug script to check what integration ranges the GUI is actually using
Add this to your batch exporter to debug the ranges issue
"""

def debug_integration_calculation(self, processor, file_info):
    """
    Debug function to check different integration methods and ranges
    """
    app_logger.info(f"\n=== DEBUGGING INTEGRATION FOR {file_info['filename']} ===")
    
    # Check if curves exist
    if not hasattr(processor, 'modified_hyperpol') or processor.modified_hyperpol is None:
        app_logger.error("No modified_hyperpol curve")
        return
    if not hasattr(processor, 'modified_depol') or processor.modified_depol is None:
        app_logger.error("No modified_depol curve")
        return
    
    app_logger.info(f"Hyperpol curve length: {len(processor.modified_hyperpol)}")
    app_logger.info(f"Depol curve length: {len(processor.modified_depol)}")
    
    # Test different integration ranges
    test_ranges = [
        (0, 200),    # Default
        (0, 100),    # Half
        (0, 50),     # Quarter  
        (50, 150),   # Middle section
        (0, len(processor.modified_hyperpol)),  # Full curve
    ]
    
    for start, end in test_ranges:
        hyperpol_integral, depol_integral = self._test_integration_range(
            processor, start, end
        )
        total = hyperpol_integral + depol_integral
        app_logger.info(f"Range {start}-{end}: Total={total:.2f}, "
                       f"Hyperpol={hyperpol_integral:.2f}, Depol={depol_integral:.2f}")
    
    # Test different time scaling
    self._test_time_scaling(processor)

def _test_integration_range(self, processor, start, end):
    """Test integration with specific range"""
    try:
        # Hyperpol
        hyperpol_len = len(processor.modified_hyperpol)
        actual_end_h = min(end, hyperpol_len)
        actual_start_h = min(start, hyperpol_len-1)
        
        if actual_start_h < actual_end_h:
            h_data = processor.modified_hyperpol[actual_start_h:actual_end_h]
            h_times = processor.modified_hyperpol_times[actual_start_h:actual_end_h]
            h_integral_raw = np.trapz(h_data, x=h_times * 1000)
            h_integral = h_integral_raw / 1000.0
        else:
            h_integral = 0.0
        
        # Depol
        depol_len = len(processor.modified_depol)
        actual_end_d = min(end, depol_len)
        actual_start_d = min(start, depol_len-1)
        
        if actual_start_d < actual_end_d:
            d_data = processor.modified_depol[actual_start_d:actual_end_d]
            d_times = processor.modified_depol_times[actual_start_d:actual_end_d]
            d_integral_raw = np.trapz(d_data, x=d_times * 1000)
            d_integral = d_integral_raw / 1000.0
        else:
            d_integral = 0.0
            
        return h_integral, d_integral
        
    except Exception as e:
        app_logger.error(f"Error in test integration: {str(e)}")
        return 0.0, 0.0

def _test_time_scaling(self, processor):
    """Test different time scaling methods"""
    app_logger.info("\n--- Testing Time Scaling ---")
    
    # Method 1: Times in seconds, convert to ms (* 1000)
    h_data = processor.modified_hyperpol[0:200]
    h_times = processor.modified_hyperpol_times[0:200]
    
    method1 = np.trapz(h_data, x=h_times * 1000) / 1000.0
    app_logger.info(f"Method 1 (times*1000/1000): {method1:.2f}")
    
    # Method 2: Times already in ms
    method2 = np.trapz(h_data, x=h_times) / 1000.0  
    app_logger.info(f"Method 2 (times as-is/1000): {method2:.2f}")
    
    # Method 3: No time conversion
    method3 = np.trapz(h_data, x=h_times * 1000)
    app_logger.info(f"Method 3 (times*1000, no /1000): {method3:.2f}")
    
    # Method 4: Direct integration without time scaling
    method4 = np.trapz(h_data, x=h_times)
    app_logger.info(f"Method 4 (direct, no scaling): {method4:.2f}")
    
    app_logger.info(f"Time range: {h_times[0]*1000:.3f} to {h_times[-1]*1000:.3f} ms")
    app_logger.info(f"Time step: {(h_times[1] - h_times[0])*1000:.3f} ms")

# ADD THIS TO YOUR _process_file_to_sheet method right before calculating integrals:

# FIXED: Calculate integrals with correct ranges and units
# ADD DEBUG CALL HERE:
self.debug_integration_calculation(processor, file_info)

hyperpol_integral, depol_integral = self._calculate_correct_integrals(processor)