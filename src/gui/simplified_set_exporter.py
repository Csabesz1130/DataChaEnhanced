"""
Simplified Set-Based Purple Curve Exporter
Complete fixed version with proper indentation and Excel file locking solution.
"""

import pandas as pd
import numpy as np
from pathlib import Path
import datetime
import re
from tkinter import filedialog, messagebox
import tkinter as tk
from tkinter import ttk
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
from src.utils.logger import app_logger
from src.io_utils.io_utils import ATFHandler
from src.analysis.action_potential import ActionPotentialProcessor
import traceback
from src.filtering.filtering import combined_filter


class SimplifiedSetExporter:
    """Simplified export manager that matches the existing Excel format exactly"""

    def __init__(self, parent_app):
        self.parent_app = parent_app

    def export_folder_by_sets(self):
        """Export ATF files organized by sets to separate Excel files"""
        try:
            # Select folder containing ATF files
            folder_path = filedialog.askdirectory(
                title="Select folder containing ATF files organized by sets"
            )

            if not folder_path:
                return False

            folder_path = Path(folder_path)

            # Find all ATF files
            atf_files = list(folder_path.glob("*.atf"))

            if not atf_files:
                messagebox.showwarning(
                    "No Files", "No ATF files found in the selected folder"
                )
                return False

            app_logger.info(f"Found {len(atf_files)} ATF files in {folder_path}")

            # Parse and organize files by sets
            file_sets = self._organize_files_by_sets(atf_files)

            if not file_sets:
                messagebox.showerror(
                    "Parse Error",
                    "Could not parse file names. Please check the naming convention.",
                )
                return False

            # Show organization confirmation
            if not self._confirm_set_organization(file_sets):
                return False

            # Select output directory
            output_dir = filedialog.askdirectory(
                title="Select output directory for Excel files"
            )

            if not output_dir:
                return False

            output_dir = Path(output_dir)

            # Process each set
            total_sets = len(file_sets)
            progress_window = self._create_progress_window(total_sets)

            try:
                processed_sets = 0
                results = []

                for set_number, files_info in file_sets.items():
                    self._update_progress(
                        progress_window,
                        processed_sets,
                        total_sets,
                        f"Processing Set {set_number}",
                    )

                    try:
                        excel_filename = self._generate_safe_excel_filename(
                            set_number, output_dir
                        )
                        success = self._create_simple_excel_for_set(
                            set_number, files_info, excel_filename
                        )

                        if success:
                            results.append(f"✓ Set {set_number}: {excel_filename.name}")
                            app_logger.info(
                                f"Set {set_number} exported to {excel_filename}"
                            )
                        else:
                            results.append(f"✗ Set {set_number}: Failed")

                    except Exception as e:
                        app_logger.error(f"Error processing set {set_number}: {str(e)}")
                        results.append(f"✗ Set {set_number}: Error - {str(e)}")

                    processed_sets += 1

                progress_window.destroy()

                # Show results
                self._show_results(results)
                return True

            finally:
                if progress_window.winfo_exists():
                    progress_window.destroy()

        except Exception as e:
            app_logger.error(f"Error in simplified set export: {str(e)}")
            messagebox.showerror(
                "Export Error", f"Failed to export file sets:\n{str(e)}"
            )
            return False

    def _organize_files_by_sets(self, atf_files):
        """Parse filenames and organize files by set number"""
        file_sets = defaultdict(list)

        # Pattern: YYYYMMDD_NNN_S_±VVmV.atf
        pattern = r"(\d{8}_\d{3})_(\d+)_([+-]?\d+)mV\.atf$"

        for file_path in atf_files:
            filename = file_path.name
            match = re.match(pattern, filename)

            if match:
                file_number = match.group(1)  # e.g., "20250528_000"
                set_number = int(match.group(2))  # e.g., 1, 2, 3
                voltage = int(match.group(3))  # e.g., -50, -10, 0

                file_info = {
                    "file_path": file_path,
                    "filename": filename,
                    "file_number": file_number,
                    "set_number": set_number,
                    "voltage": voltage,
                    "sheet_name": file_number,
                }

                file_sets[set_number].append(file_info)

            else:
                app_logger.warning(f"Could not parse filename: {filename}")

        # Sort files within each set by file number
        for set_number in file_sets:
            file_sets[set_number].sort(key=lambda x: x["file_number"])

        return dict(file_sets)

    def _confirm_set_organization(self, file_sets):
        """Show user the set organization and confirm"""
        confirm_window = tk.Toplevel(self.parent_app.master)
        confirm_window.title("Confirm Set Organization")
        confirm_window.transient(self.parent_app.master)
        confirm_window.grab_set()

        # Center window with minimum size
        confirm_window.update_idletasks()
        desired_width = 600
        desired_height = 400
        x = (confirm_window.winfo_screenwidth() // 2) - (desired_width // 2)
        y = (confirm_window.winfo_screenheight() // 2) - (desired_height // 2)
        confirm_window.geometry(f"{desired_width}x{desired_height}+{x}+{y}")
        confirm_window.minsize(desired_width, desired_height)

        # Title
        tk.Label(
            confirm_window, text="File Set Organization", font=("Arial", 14, "bold")
        ).pack(pady=10)

        # Scrollable text area
        frame = tk.Frame(confirm_window)
        frame.pack(fill="both", expand=True, padx=20, pady=10)

        scrollbar = tk.Scrollbar(frame)
        scrollbar.pack(side="right", fill="y")

        text_area = tk.Text(frame, yscrollcommand=scrollbar.set, wrap="word")
        text_area.pack(side="left", fill="both", expand=True)
        scrollbar.config(command=text_area.yview)

        # Build organization text
        org_text = f"Found {len(file_sets)} sets:\n\n"

        for set_number in sorted(file_sets.keys()):
            files_info = file_sets[set_number]
            org_text += f"📁 SET {set_number} ({len(files_info)} files) → 'Set_{set_number}_[timestamp].xlsx'\n"

            for file_info in files_info:
                org_text += f"   📄 {file_info['filename']} → Sheet: '{file_info['sheet_name']}'\n"

            org_text += "\n"

        text_area.insert("1.0", org_text)
        text_area.config(state="disabled")

        # Result variable
        result = {"confirmed": False}

        # Buttons
        button_frame = tk.Frame(confirm_window)
        button_frame.pack(pady=10)

        def confirm():
            result["confirmed"] = True
            confirm_window.destroy()

        def cancel():
            result["confirmed"] = False
            confirm_window.destroy()

        tk.Button(
            button_frame,
            text="✓ Export",
            command=confirm,
            bg="green",
            fg="white",
            font=("Arial", 10, "bold"),
        ).pack(side="left", padx=10)
        tk.Button(
            button_frame,
            text="✗ Cancel",
            command=cancel,
            bg="red",
            fg="white",
            font=("Arial", 10, "bold"),
        ).pack(side="left", padx=10)

        confirm_window.wait_window()
        return result["confirmed"]

    def _generate_safe_excel_filename(self, set_number, output_dir):
        """Generate Excel filename that avoids conflicts with open files"""
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")

        # Try different filename variations to avoid locked files
        for attempt in range(10):
            if attempt == 0:
                filename = f"Set_{set_number}_{timestamp}.xlsx"
            else:
                filename = f"Set_{set_number}_{timestamp}_v{attempt}.xlsx"

            full_path = output_dir / filename

            # Test if we can create/write to this file
            if self._test_file_writable(full_path):
                app_logger.info(f"Generated safe filename: {filename}")
                return full_path

        # If all attempts fail, ask user what to do
        app_logger.warning("Could not generate writable filename automatically")
        return self._handle_filename_conflict(set_number, output_dir)

    def _test_file_writable(self, file_path):
        """Test if a file path is writable without actually creating the file"""
        try:
            # Try to open in write mode
            with open(file_path, "w") as f:
                pass
            # If successful, remove the test file
            file_path.unlink()
            return True
        except (OSError, PermissionError):
            return False

    def _handle_filename_conflict(self, set_number, output_dir):
        """Handle filename conflicts by asking user for guidance"""
        conflict_window = tk.Toplevel(self.parent_app.master)
        conflict_window.title("File Conflict")
        conflict_window.transient(self.parent_app.master)
        conflict_window.grab_set()

        # Center window
        conflict_window.update_idletasks()
        desired_width = 500
        desired_height = 200
        x = (conflict_window.winfo_screenwidth() // 2) - (desired_width // 2)
        y = (conflict_window.winfo_screenheight() // 2) - (desired_height // 2)
        conflict_window.geometry(f"{desired_width}x{desired_height}+{x}+{y}")
        conflict_window.minsize(desired_width, desired_height)

        # Message
        message = (
            f"Cannot create Excel file for Set {set_number}.\n"
            f"Possible causes:\n"
            f"• Excel files are open in another application\n"
            f"• Output directory is write-protected\n\n"
            f"Please close Excel files or choose a different output directory."
        )

        tk.Label(conflict_window, text=message, wraplength=450, justify="left").pack(
            pady=20
        )

        result = {"action": "skip"}

        button_frame = tk.Frame(conflict_window)
        button_frame.pack(pady=10)

        def retry():
            result["action"] = "retry"
            conflict_window.destroy()

        def skip():
            result["action"] = "skip"
            conflict_window.destroy()

        tk.Button(
            button_frame, text="Retry", command=retry, bg="blue", fg="white"
        ).pack(side="left", padx=10)
        tk.Button(
            button_frame, text="Skip Set", command=skip, bg="orange", fg="white"
        ).pack(side="left", padx=10)

        conflict_window.wait_window()

        if result["action"] == "retry":
            return self._generate_safe_excel_filename(set_number, output_dir)
        else:
            # Return a placeholder that will cause the set to be skipped
            return None

    def _create_simple_excel_for_set(self, set_number, files_info, excel_filename):
        """Create simplified Excel file matching the existing format"""
        if excel_filename is None:
            app_logger.warning(f"Skipping Set {set_number} due to filename conflict")
            return False

        try:
            app_logger.info(
                f"Creating simplified Excel for Set {set_number}: {excel_filename}"
            )

            wb = Workbook()

            # Process each file in the set
            processed_files = []

            for file_info in files_info:
                try:
                    # Process the ATF file
                    processor = self._process_atf_file_simple(file_info)

                    if processor is not None:
                        # Add simple sheet to workbook
                        self._add_simple_sheet(wb, file_info, processor)
                        processed_files.append(file_info)

                except Exception as e:
                    app_logger.error(
                        f"Error processing {file_info['filename']}: {str(e)}"
                    )
                    continue

            # Remove default sheet if we have data sheets
            if processed_files and "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])

            # Save workbook with final safety check
            try:
                wb.save(excel_filename)
                app_logger.info(
                    f"Set {set_number} saved: {len(processed_files)} files processed"
                )
                return True
            except PermissionError as e:
                app_logger.error(f"Permission denied saving {excel_filename}: {str(e)}")
                return False

        except Exception as e:
            app_logger.error(
                f"Error creating simplified Excel for set {set_number}: {str(e)}"
            )
            return False

        # Ensure this import is at the top of your src/gui/simplified_set_exporter.py file
        # from src.filtering.filtering import combined_filter
        # import numpy as np
        # from openpyxl.styles import Font
        # from openpyxl.utils import get_column_letter
        # import traceback

        # --- Insert this function into your SimplifiedSetExporter class ---

    def _process_atf_file_simple(self, file_info):
        """
        Process ATF file following the GUI workflow:
        1. Load file
        2. Apply specified filters (Savitzky-Golay and Butterworth) 
        3. Set processor parameters including cycle definitions and "Averaged Normalized" method
        4. Generate purple curves.
        """
        file_path = file_info['file_path']
        filename_short = file_info['filename']
        voltage = file_info['voltage']
        
        app_logger.info(f"Processing ATF file: {filename_short} with V2={voltage}mV following GUI workflow (Corrected v6)")

        try:
            atf_handler = ATFHandler(file_path)
            atf_handler.load_atf()
            
            time_data = atf_handler.get_column("Time")
            raw_data = atf_handler.get_column("#1")

            if time_data is None or raw_data is None or len(time_data) == 0 or len(raw_data) == 0 or len(time_data) != len(raw_data):
                app_logger.error(f"Data loading issue for {filename_short}. Time: {len(time_data) if time_data is not None else 'None'}, Data: {len(raw_data) if raw_data is not None else 'None'}")
                return None
            app_logger.debug(f"Loaded {len(raw_data)} raw data points for {filename_short}.")

            sampling_interval = atf_handler.get_sampling_rate()
            if sampling_interval <= 0:
                app_logger.warning(f"Invalid sampling interval ({sampling_interval}s) for {filename_short}. Defaulting fs to 10kHz.")
                fs_hz = 10000.0 
            else:
                fs_hz = 1.0 / sampling_interval
            app_logger.debug(f"Calculated sampling frequency for {filename_short}: {fs_hz:.2f} Hz")

            # Filter parameters based on user's GUI settings (image_beb633.png)
            # Savitzky-Golay: Window 101, Polyorder 3
            # Butterworth: Cutoff 2000 Hz, Order 2
            filter_params_for_export = {
                'savgol_params': {'window_length': 101, 'polyorder': 3},
                'butter_params': {'cutoff': 2000, 'fs': fs_hz, 'order': 2}
            }
            app_logger.info(f"Applying GUI-matched filters to {filename_short}: {filter_params_for_export}")
            
            filtered_data_for_export = combined_filter(raw_data, **filter_params_for_export)
            app_logger.debug(f"Data filtered for {filename_short}. Length: {len(filtered_data_for_export)}")

            # Parameters for ActionPotentialProcessor
            # Includes cycle parameters and "Averaged Normalized" method equivalent
            processor_params = {
                'V0': -80, 
                'V2': voltage,
                'time_constant_ms': 10, 
                'threshold_percent': 10,
                'use_alternative_method': True, # For "Averaged Normalized" logic in process_signal
                'show_modified_peaks': True,
                # Explicitly provide cycle parameters that ActionPotentialProcessor.process_signal expects
                # (these were previously causing KeyErrors or "No cycles found")
                'n_cycles': 2,  # Default or typical value
                't0': 20,       # Default or typical value (ms)
                't1': 100,      # Default or typical value (ms)
                't2': 100,      # Default or typical value (ms)
                't3': 1000,     # Default or typical value (ms)
                # Add other parameters if your ActionPotentialProcessor's __init__ or process_signal needs them
                # For example, if these were in your earlier "Parameters validated" log and are used:
                # 'V1': -100, 
                # 'cell_area_cm2': 0.0001,
                # 'window_size': 51, # Usually for internal SavGol in processor, might not be needed if input is pre-filtered
                # 'polyorder': 3,    # Usually for internal SavGol in processor
                'purple_curve_duration_ms': 99.5, # To aim for 199 intervals (200 points @ 0.5ms step for linspace)
                                                  # Or (199 points * 0.5ms step = 99.5ms span)
                'purple_curve_num_points': 199,   # Explicitly target 199 points
            }
            
            processor = ActionPotentialProcessor(data=filtered_data_for_export, 
                                                 time_data=time_data, 
                                                 params=processor_params)
            
            app_logger.debug(f"Processor params for {filename_short} upon instantiation: {processor.params}")
            
            # --- Analysis Sequence ---
            # 1. Call process_signal (use_alternative_method is set via params)
            _p_data, _o_curve, _o_times, _n_curve, _n_times, _avg_curve, _avg_times, results_signal = processor.process_signal()

            if results_signal.get('error'):
                error_message = results_signal['error']
                app_logger.error(f"process_signal failed for {filename_short}: {error_message}")
                if "No cycles found" in str(error_message) or any(key in str(error_message) for key in ['t0','t1','t2','t3','n_cycles']):
                    app_logger.warning(
                        f"Consider adjusting cycle parameters (t0-t3, n_cycles) in SimplifiedSetExporter for file {filename_short}. "
                        f"Current cycle params used: n_cycles={processor.params.get('n_cycles')}, "
                        f"t0={processor.params.get('t0')}, t1={processor.params.get('t1')}, "
                        f"t2={processor.params.get('t2')}, t3={processor.params.get('t3')}"
                    )
                return None
            app_logger.debug(f"process_signal completed for {filename_short}.")

            # 2. Call apply_average_to_peaks (NO ARGUMENTS)
            processor.apply_average_to_peaks()
            app_logger.debug(f"apply_average_to_peaks called for {filename_short}.")
            
            if (not hasattr(processor, 'modified_hyperpol') or processor.modified_hyperpol is None or
                not hasattr(processor, 'modified_depol') or processor.modified_depol is None or
                not hasattr(processor, 'modified_hyperpol_times') or processor.modified_hyperpol_times is None or
                not hasattr(processor, 'modified_depol_times') or processor.modified_depol_times is None):
                app_logger.error(f"Purple curve data attributes not found for {filename_short} after analysis steps.")
                return None

            expected_purple_points = 199 
            hyperpol_len = len(processor.modified_hyperpol)
            depol_len = len(processor.modified_depol)

            if hyperpol_len != expected_purple_points or depol_len != expected_purple_points:
                app_logger.error(f"Purple curves data arrays for {filename_short} have incorrect length. "
                                 f"Expected: {expected_purple_points}, Got Hyperpol: {hyperpol_len}, Depol: {depol_len}")
                return None
            
            # Also check times arrays length
            if len(processor.modified_hyperpol_times) != expected_purple_points or \
               len(processor.modified_depol_times) != expected_purple_points:
                app_logger.error(f"Purple curve times arrays for {filename_short} have incorrect length. "
                                 f"Expected: {expected_purple_points}, "
                                 f"Got Hyperpol_times: {len(processor.modified_hyperpol_times)}, "
                                 f"Depol_times: {len(processor.modified_depol_times)}")
                return None

            processor.analysis_complete = True 
            app_logger.info(f"Successfully processed {filename_short}. Purple curves generated with {expected_purple_points} points.")
            return processor
            
        except Exception as e:
            app_logger.error(f"Exception in _process_atf_file_simple for {filename_short}: {str(e)}\n{traceback.format_exc()}")
            return None

    def _add_simple_sheet(self, wb, file_info, processor):
        sheet_name_base = file_info['sheet_name']
        sheet_name = sheet_name_base
        counter = 1
        while sheet_name in wb.sheetnames:
            sheet_name = f"{sheet_name_base}_{counter}"
            counter += 1
        ws = wb.create_sheet(title=sheet_name)
        app_logger.info(f"Adding sheet '{ws.title}' for file {file_info['filename']}")

        try:
            expected_purple_points = 199
            # Comprehensive check for all required attributes and their lengths
            if not (hasattr(processor, 'modified_hyperpol') and isinstance(processor.modified_hyperpol, np.ndarray) and
                    hasattr(processor, 'modified_depol') and isinstance(processor.modified_depol, np.ndarray) and
                    hasattr(processor, 'modified_hyperpol_times') and isinstance(processor.modified_hyperpol_times, np.ndarray) and
                    hasattr(processor, 'modified_depol_times') and isinstance(processor.modified_depol_times, np.ndarray) and
                    len(processor.modified_hyperpol) == expected_purple_points and
                    len(processor.modified_depol) == expected_purple_points and 
                    len(processor.modified_hyperpol_times) == expected_purple_points and
                    len(processor.modified_depol_times) == expected_purple_points):
                app_logger.error(f"Processor for sheet '{ws.title}' is missing critical purple curve data, data is not {expected_purple_points} points long, or types are incorrect, or times arrays mismatch.")
                ws['A1'] = f"Error: Missing, incomplete, or mismatched processed data for {file_info['filename']} (expected {expected_purple_points} points)."
                return

            integration_end_idx = expected_purple_points 
            hyperpol_integral = processor.calculate_integration('hyperpol', 0, integration_end_idx)
            depol_integral = processor.calculate_integration('depol', 0, integration_end_idx)
            overall_integral = hyperpol_integral + depol_integral 

            app_logger.debug(f"Sheet '{ws.title}': Hyperpol Integral={hyperpol_integral:.3f}, Depol Integral={depol_integral:.3f}, Overall={overall_integral:.3f}")

            current_row = 1
            
            ws[f'A{current_row}'] = "Original ATF File:"
            ws[f'B{current_row}'] = file_info['filename']
            current_row += 1
            ws[f'A{current_row}'] = "V2 Voltage (mV):"
            ws[f'B{current_row}'] = file_info['voltage']
            current_row += 2

            ws[f'A{current_row}'] = "Overall Integral (pC):"
            ws[f'B{current_row}'] = float(f"{overall_integral:.2f}")
            ws[f'A{current_row}'].font = Font(bold=True)
            current_row += 2
            
            ws[f'A{current_row}'] = "PURPLE HYPERPOL CURVE"
            ws[f'A{current_row}'].font = Font(bold=True)
            current_row += 1
            
            ws[f'A{current_row}'] = "Hyperpol Integral (pC):"
            ws[f'B{current_row}'] = float(f"{hyperpol_integral:.2f}")
            current_row += 1
            
            headers_hyperpol = ["Index", "Hyperpol_pA", "Hyperpol_time_ms"]
            for col_idx, header_title in enumerate(headers_hyperpol):
                col_letter = get_column_letter(col_idx + 1)
                ws[f'{col_letter}{current_row}'] = header_title
                ws[f'{col_letter}{current_row}'].font = Font(bold=True)
            current_row += 1
            
            for i in range(expected_purple_points):
                ws[f'A{current_row + i}'] = int(i + 1)
                ws[f'B{current_row + i}'] = float(f"{processor.modified_hyperpol[i]:.7f}")
                ws[f'C{current_row + i}'] = float(f"{processor.modified_hyperpol_times[i] * 1000:.7f}")
            current_row += expected_purple_points + 2 
            
            ws[f'A{current_row}'] = "PURPLE DEPOL CURVE"
            ws[f'A{current_row}'].font = Font(bold=True)
            current_row += 1

            ws[f'A{current_row}'] = "Depol Integral (pC):"
            ws[f'B{current_row}'] = float(f"{depol_integral:.2f}")
            current_row += 1
            
            headers_depol = ["Index", "Depol_pA", "Depol_time_ms"]
            for col_idx, header_title in enumerate(headers_depol):
                col_letter = get_column_letter(col_idx + 1)
                ws[f'{col_letter}{current_row}'] = header_title
                ws[f'{col_letter}{current_row}'].font = Font(bold=True)
            current_row += 1
            
            for i in range(expected_purple_points):
                ws[f'A{current_row + i}'] = int(i + 1)
                ws[f'B{current_row + i}'] = float(f"{processor.modified_depol[i]:.7f}")
                ws[f'C{current_row + i}'] = float(f"{processor.modified_depol_times[i] * 1000:.7f}")
            
            for col_cells in ws.columns:
                max_length = 0
                column_letter_str = get_column_letter(col_cells[0].column)
                for cell in col_cells:
                    try:
                        if cell.value is not None:
                            cell_len = len(str(cell.value))
                            if cell_len > max_length:
                                max_length = cell_len
                    except:
                        pass
                adjusted_width = (max_length + 2) if max_length > 0 else 10
                ws.column_dimensions[column_letter_str].width = adjusted_width

            app_logger.info(f"Successfully populated sheet '{ws.title}' for {file_info['filename']}")

        except Exception as e:
            app_logger.error(f"Error populating sheet '{ws.title}' for {file_info['filename']}: {str(e)}\n{traceback.format_exc()}")
            try:
                ws['A1'] = f"Error populating this sheet for file: {file_info['filename']}"
                ws['A2'] = str(e)
            except:
                pass

    def _create_progress_window(self, total_sets):
        """Create progress window"""
        progress_window = tk.Toplevel(self.parent_app.master)
        progress_window.title("Processing Sets")
        progress_window.transient(self.parent_app.master)
        progress_window.grab_set()

        # Center window with minimum size
        progress_window.update_idletasks()
        desired_width = 450
        desired_height = 120
        x = (progress_window.winfo_screenwidth() // 2) - (desired_width // 2)
        y = (progress_window.winfo_screenheight() // 2) - (desired_height // 2)
        progress_window.geometry(f"{desired_width}x{desired_height}+{x}+{y}")
        progress_window.minsize(desired_width, desired_height)

        # Progress elements
        progress_window.label = tk.Label(progress_window, text="Starting processing...")
        progress_window.label.pack(pady=10)

        progress_window.progress = ttk.Progressbar(
            progress_window, length=350, mode="determinate", maximum=total_sets
        )
        progress_window.progress.pack(pady=5)

        progress_window.status = tk.Label(progress_window, text="")
        progress_window.status.pack(pady=5)

        progress_window.update()
        return progress_window

    def _update_progress(self, progress_window, current, total, status=""):
        """Update progress window"""
        try:
            if progress_window.winfo_exists():
                progress_window.progress["value"] = current
                progress_window.label.config(
                    text=f"Processing set {current + 1}/{total}"
                )
                progress_window.status.config(text=status)
                progress_window.update()
        except:
            pass

    def _show_results(self, results):
        """Show export results to user"""
        result_window = tk.Toplevel(self.parent_app.master)
        result_window.title("Export Results")
        result_window.transient(self.parent_app.master)

        # Center window with minimum size
        result_window.update_idletasks()
        desired_width = 500
        desired_height = 300
        x = (result_window.winfo_screenwidth() // 2) - (desired_width // 2)
        y = (result_window.winfo_screenheight() // 2) - (desired_height // 2)
        result_window.geometry(f"{desired_width}x{desired_height}+{x}+{y}")
        result_window.minsize(desired_width, desired_height)

        tk.Label(result_window, text="Export Results", font=("Arial", 12, "bold")).pack(
            pady=10
        )

        # Scrollable results
        frame = tk.Frame(result_window)
        frame.pack(fill="both", expand=True, padx=20, pady=10)

        scrollbar = tk.Scrollbar(frame)
        scrollbar.pack(side="right", fill="y")

        text_area = tk.Text(frame, yscrollcommand=scrollbar.set)
        text_area.pack(side="left", fill="both", expand=True)
        scrollbar.config(command=text_area.yview)

        results_text = "\n".join(results)
        text_area.insert("1.0", results_text)
        text_area.config(state="disabled")

        tk.Button(result_window, text="Close", command=result_window.destroy).pack(
            pady=10
        )


# Integration with main application
class SimplifiedSetExportButton:
    """Simplified set export button for integration with the main app"""

    def __init__(self, parent_app, toolbar_frame):
        self.parent_app = parent_app
        self.exporter = SimplifiedSetExporter(parent_app)

        # Add export button to toolbar
        self.export_button = tk.Button(
            toolbar_frame,
            text="📊 Export Sets",
            command=self.export_sets,
            relief="raised",
            bg="#4A90E2",
            fg="white",
            font=("Arial", 9, "bold"),
            padx=8,
            pady=2,
        )
        self.export_button.pack(side="left", padx=2)

        app_logger.info("Simplified set export button added to toolbar")

    def export_sets(self):
        """Export ATF file sets with simplified format"""
        try:
            self.exporter.export_folder_by_sets()
        except Exception as e:
            app_logger.error(f"Error in set export: {str(e)}")
            messagebox.showerror("Export Error", f"Failed to export sets:\n{str(e)}")


# Standalone function for direct integration
def add_set_export_to_toolbar(parent_app, toolbar_frame):
    """
    Add the set export button to an existing toolbar.

    Args:
        parent_app: The main application instance
        toolbar_frame: The tkinter frame where the button should be added

    Returns:
        SimplifiedSetExportButton: The created button instance
    """
    return SimplifiedSetExportButton(parent_app, toolbar_frame)
