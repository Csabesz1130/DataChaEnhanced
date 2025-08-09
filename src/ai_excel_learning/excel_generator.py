"""
Excel Generator for AI Learning

This module generates complete Excel files with data, charts, and formatting
based on learned patterns and models.
"""

import pandas as pd
import numpy as np
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Alignment, NamedStyle
from openpyxl.chart import ScatterChart, LineChart, BarChart, PieChart
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension, RowDimension
import json
import logging
from typing import Dict, List, Tuple, Any, Optional
from dataclasses import dataclass
from pathlib import Path
from datetime import datetime

logger = logging.getLogger(__name__)

@dataclass
class ExcelGenerationConfig:
    """Configuration for Excel file generation"""
    filename: str
    sheets: List[Dict[str, Any]]
    charts: List[Dict[str, Any]]
    formatting: Dict[str, Any]
    metadata: Dict[str, Any]

class ExcelGenerator:
    """
    Generates complete Excel files based on learned patterns
    """
    
    def __init__(self):
        self.generation_history = []
        
    def generate_excel_file(self, data: pd.DataFrame, 
                           config: ExcelGenerationConfig,
                           output_path: str) -> str:
        """
        Generate a complete Excel file
        
        Args:
            data: Data to include in the file
            config: Generation configuration
            output_path: Path to save the file
            
        Returns:
            Path to the generated file
        """
        logger.info(f"Generating Excel file: {output_path}")
        
        # Create workbook
        workbook = Workbook()
        
        # Remove default sheet
        workbook.remove(workbook.active)
        
        # Generate sheets
        for sheet_config in config.sheets:
            self._generate_sheet(workbook, data, sheet_config)
        
        # Generate charts
        for chart_config in config.charts:
            self._generate_chart(workbook, chart_config)
        
        # Apply formatting
        self._apply_formatting(workbook, config.formatting)
        
        # Add metadata
        self._add_metadata(workbook, config.metadata)
        
        # Save file
        workbook.save(output_path)
        
        # Record generation
        self.generation_history.append({
            'timestamp': datetime.now().isoformat(),
            'output_path': output_path,
            'config': config.__dict__,
            'data_shape': data.shape
        })
        
        logger.info(f"Excel file generated successfully: {output_path}")
        return output_path
    
    def _generate_sheet(self, workbook: Workbook, data: pd.DataFrame, 
                       sheet_config: Dict[str, Any]):
        """Generate a single sheet"""
        sheet_name = sheet_config.get('name', 'Sheet1')
        sheet = workbook.create_sheet(title=sheet_name)
        
        # Add data to sheet
        self._add_data_to_sheet(sheet, data, sheet_config)
        
        # Apply sheet-specific formatting
        if 'formatting' in sheet_config:
            self._apply_sheet_formatting(sheet, sheet_config['formatting'])
        
        # Add formulas if specified
        if 'formulas' in sheet_config:
            self._add_formulas_to_sheet(sheet, sheet_config['formulas'])
    
    def _add_data_to_sheet(self, sheet, data: pd.DataFrame, sheet_config: Dict[str, Any]):
        """Add data to a sheet"""
        start_row = sheet_config.get('start_row', 1)
        start_col = sheet_config.get('start_col', 1)
        
        # Add headers if specified
        if sheet_config.get('include_headers', True):
            for i, col_name in enumerate(data.columns):
                cell = sheet.cell(row=start_row, column=start_col + i)
                cell.value = col_name
                cell.font = Font(bold=True)
            
            start_row += 1
        
        # Add data
        for i, row in enumerate(data.itertuples(index=False)):
            for j, value in enumerate(row):
                cell = sheet.cell(row=start_row + i, column=start_col + j)
                cell.value = value
    
    def _apply_sheet_formatting(self, sheet, formatting: Dict[str, Any]):
        """Apply formatting to a sheet"""
        # Column formatting
        for col_config in formatting.get('columns', []):
            col_letter = col_config.get('column', 'A')
            width = col_config.get('width')
            if width:
                sheet.column_dimensions[col_letter].width = width
        
        # Row formatting
        for row_config in formatting.get('rows', []):
            row_num = row_config.get('row', 1)
            height = row_config.get('height')
            if height:
                sheet.row_dimensions[row_num].height = height
        
        # Cell formatting
        for cell_config in formatting.get('cells', []):
            cell_ref = cell_config.get('cell', 'A1')
            cell = sheet[cell_ref]
            
            # Font formatting
            if 'font' in cell_config:
                font_config = cell_config['font']
                cell.font = Font(
                    name=font_config.get('name', 'Calibri'),
                    size=font_config.get('size', 11),
                    bold=font_config.get('bold', False),
                    italic=font_config.get('italic', False),
                    color=font_config.get('color')
                )
            
            # Fill formatting
            if 'fill' in cell_config:
                fill_config = cell_config['fill']
                cell.fill = PatternFill(
                    start_color=fill_config.get('start_color'),
                    end_color=fill_config.get('end_color'),
                    fill_type=fill_config.get('fill_type', 'solid')
                )
            
            # Border formatting
            if 'border' in cell_config:
                border_config = cell_config['border']
                cell.border = Border(
                    left=border_config.get('left'),
                    right=border_config.get('right'),
                    top=border_config.get('top'),
                    bottom=border_config.get('bottom')
                )
            
            # Alignment formatting
            if 'alignment' in cell_config:
                align_config = cell_config['alignment']
                cell.alignment = Alignment(
                    horizontal=align_config.get('horizontal', 'general'),
                    vertical=align_config.get('vertical', 'bottom'),
                    wrap_text=align_config.get('wrap_text', False)
                )
    
    def _add_formulas_to_sheet(self, sheet, formulas: List[Dict[str, Any]]):
        """Add formulas to a sheet"""
        for formula_config in formulas:
            cell_ref = formula_config.get('cell', 'A1')
            formula = formula_config.get('formula', '=SUM(A1:A10)')
            
            cell = sheet[cell_ref]
            cell.value = formula
    
    def _generate_chart(self, workbook: Workbook, chart_config: Dict[str, Any]):
        """Generate a chart"""
        sheet_name = chart_config.get('sheet', 'Sheet1')
        chart_type = chart_config.get('type', 'scatter')
        
        if sheet_name not in workbook.sheetnames:
            logger.warning(f"Sheet {sheet_name} not found for chart generation")
            return
        
        sheet = workbook[sheet_name]
        
        # Create chart based on type
        if chart_type == 'scatter':
            chart = self._create_scatter_chart(chart_config)
        elif chart_type == 'line':
            chart = self._create_line_chart(chart_config)
        elif chart_type == 'bar':
            chart = self._create_bar_chart(chart_config)
        elif chart_type == 'pie':
            chart = self._create_pie_chart(chart_config)
        else:
            logger.warning(f"Unsupported chart type: {chart_type}")
            return
        
        # Add chart to sheet
        chart_position = chart_config.get('position', {})
        sheet.add_chart(chart, chart_position.get('anchor', 'A1'))
    
    def _create_scatter_chart(self, chart_config: Dict[str, Any]) -> ScatterChart:
        """Create a scatter chart"""
        chart = ScatterChart()
        
        # Set chart title
        if 'title' in chart_config:
            chart.title = chart_config['title']
        
        # Configure axes
        if 'x_axis' in chart_config:
            x_config = chart_config['x_axis']
            chart.x_axis.title = x_config.get('title', 'X Axis')
            if 'min' in x_config:
                chart.x_axis.min = x_config['min']
            if 'max' in x_config:
                chart.x_axis.max = x_config['max']
        
        if 'y_axis' in chart_config:
            y_config = chart_config['y_axis']
            chart.y_axis.title = y_config.get('title', 'Y Axis')
            if 'min' in y_config:
                chart.y_axis.min = y_config['min']
            if 'max' in y_config:
                chart.y_axis.max = y_config['max']
        
        # Add series
        for series_config in chart_config.get('series', []):
            series = chart.add_series(
                xvalues=series_config.get('x_values', 'A1:A10'),
                yvalues=series_config.get('y_values', 'B1:B10'),
                title=series_config.get('title', 'Series')
            )
            
            # Configure series properties
            if 'marker' in series_config:
                series.marker.symbol = series_config['marker']
        
        return chart
    
    def _create_line_chart(self, chart_config: Dict[str, Any]) -> LineChart:
        """Create a line chart"""
        chart = LineChart()
        
        # Set chart title
        if 'title' in chart_config:
            chart.title = chart_config['title']
        
        # Configure axes
        if 'x_axis' in chart_config:
            x_config = chart_config['x_axis']
            chart.x_axis.title = x_config.get('title', 'X Axis')
        
        if 'y_axis' in chart_config:
            y_config = chart_config['y_axis']
            chart.y_axis.title = y_config.get('title', 'Y Axis')
        
        # Add series
        for series_config in chart_config.get('series', []):
            series = chart.add_series(
                xvalues=series_config.get('x_values', 'A1:A10'),
                yvalues=series_config.get('y_values', 'B1:B10'),
                title=series_config.get('title', 'Series')
            )
            
            # Configure line properties
            if 'line_style' in series_config:
                series.line.style = series_config['line_style']
        
        return chart
    
    def _create_bar_chart(self, chart_config: Dict[str, Any]) -> BarChart:
        """Create a bar chart"""
        chart = BarChart()
        
        # Set chart title
        if 'title' in chart_config:
            chart.title = chart_config['title']
        
        # Configure axes
        if 'x_axis' in chart_config:
            x_config = chart_config['x_axis']
            chart.x_axis.title = x_config.get('title', 'Categories')
        
        if 'y_axis' in chart_config:
            y_config = chart_config['y_axis']
            chart.y_axis.title = y_config.get('title', 'Values')
        
        # Add series
        for series_config in chart_config.get('series', []):
            series = chart.add_series(
                xvalues=series_config.get('x_values', 'A1:A10'),
                yvalues=series_config.get('y_values', 'B1:B10'),
                title=series_config.get('title', 'Series')
            )
        
        return chart
    
    def _create_pie_chart(self, chart_config: Dict[str, Any]) -> PieChart:
        """Create a pie chart"""
        chart = PieChart()
        
        # Set chart title
        if 'title' in chart_config:
            chart.title = chart_config['title']
        
        # Add series
        for series_config in chart_config.get('series', []):
            series = chart.add_series(
                xvalues=series_config.get('x_values', 'A1:A10'),
                yvalues=series_config.get('y_values', 'B1:B10'),
                title=series_config.get('title', 'Series')
            )
        
        return chart
    
    def _apply_formatting(self, workbook: Workbook, formatting: Dict[str, Any]):
        """Apply global formatting to workbook"""
        # Apply to all sheets
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # Apply default styles
            if 'default_styles' in formatting:
                self._apply_default_styles(sheet, formatting['default_styles'])
            
            # Apply conditional formatting
            if 'conditional_formatting' in formatting:
                self._apply_conditional_formatting(sheet, formatting['conditional_formatting'])
    
    def _apply_default_styles(self, sheet, styles: Dict[str, Any]):
        """Apply default styles to a sheet"""
        # Header style
        if 'header_style' in styles:
            header_style = styles['header_style']
            for cell in sheet[1]:
                cell.font = Font(
                    bold=header_style.get('bold', True),
                    size=header_style.get('size', 12)
                )
        
        # Data style
        if 'data_style' in styles:
            data_style = styles['data_style']
            for row in sheet.iter_rows(min_row=2):
                for cell in row:
                    cell.font = Font(
                        size=data_style.get('size', 11)
                    )
    
    def _apply_conditional_formatting(self, sheet, conditional_formats: List[Dict[str, Any]]):
        """Apply conditional formatting to a sheet"""
        for format_config in conditional_formats:
            range_ref = format_config.get('range', 'A1:Z100')
            rule_type = format_config.get('type', 'cellIs')
            
            if rule_type == 'cellIs':
                # Simple cell value rule
                operator = format_config.get('operator', 'greaterThan')
                value = format_config.get('value', 0)
                
                # This is a simplified implementation
                # In production, you'd use openpyxl's conditional formatting features
                pass
    
    def _add_metadata(self, workbook: Workbook, metadata: Dict[str, Any]):
        """Add metadata to workbook"""
        # Properties
        if 'properties' in metadata:
            props = metadata['properties']
            workbook.properties.title = props.get('title', 'Generated Excel File')
            workbook.properties.subject = props.get('subject', 'AI Generated')
            workbook.properties.creator = props.get('creator', 'DataChaEnhanced AI')
            workbook.properties.created = props.get('created', datetime.now())
        
        # Defined names
        if 'defined_names' in metadata:
            for name_config in metadata['defined_names']:
                workbook.define_name(
                    name_config['name'],
                    name_config['value']
                )
    
    def generate_from_template(self, template_path: str, data: pd.DataFrame, 
                              output_path: str) -> str:
        """
        Generate Excel file from a template
        
        Args:
            template_path: Path to template file
            data: Data to include
            output_path: Path to save the file
            
        Returns:
            Path to the generated file
        """
        # Load template
        with open(template_path, 'r') as f:
            template_config = json.load(f)
        
        # Create generation config from template
        config = ExcelGenerationConfig(**template_config)
        
        # Generate file
        return self.generate_excel_file(data, config, output_path)
    
    def create_template_from_existing(self, excel_path: str, template_path: str):
        """
        Create a template from an existing Excel file
        
        Args:
            excel_path: Path to existing Excel file
            template_path: Path to save template
            
        Returns:
            Template configuration
        """
        logger.info(f"Creating template from: {excel_path}")
        
        workbook = openpyxl.load_workbook(excel_path)
        
        template_config = {
            'filename': Path(excel_path).stem,
            'sheets': [],
            'charts': [],
            'formatting': {},
            'metadata': {}
        }
        
        # Extract sheet configurations
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_config = self._extract_sheet_config(sheet, sheet_name)
            template_config['sheets'].append(sheet_config)
        
        # Extract chart configurations
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for chart in sheet._charts:
                chart_config = self._extract_chart_config(chart, sheet_name)
                template_config['charts'].append(chart_config)
        
        # Extract formatting
        template_config['formatting'] = self._extract_formatting_config(workbook)
        
        # Extract metadata
        template_config['metadata'] = self._extract_metadata_config(workbook)
        
        # Save template
        with open(template_path, 'w') as f:
            json.dump(template_config, f, indent=2)
        
        logger.info(f"Template saved to: {template_path}")
        return template_config
    
    def _extract_sheet_config(self, sheet, sheet_name: str) -> Dict[str, Any]:
        """Extract configuration from a sheet"""
        config = {
            'name': sheet_name,
            'start_row': 1,
            'start_col': 1,
            'include_headers': True,
            'formatting': {}
        }
        
        # Extract column widths
        column_widths = []
        for col in sheet.column_dimensions:
            width = sheet.column_dimensions[col].width
            if width:
                column_widths.append({
                    'column': col,
                    'width': width
                })
        
        if column_widths:
            config['formatting']['columns'] = column_widths
        
        return config
    
    def _extract_chart_config(self, chart, sheet_name: str) -> Dict[str, Any]:
        """Extract configuration from a chart"""
        config = {
            'sheet': sheet_name,
            'type': type(chart).__name__.lower().replace('chart', ''),
            'series': []
        }
        
        # Extract chart title
        if hasattr(chart, 'title') and chart.title:
            config['title'] = str(chart.title)
        
        # Extract series
        for series in chart.series:
            series_config = {
                'title': series.title or 'Series',
                'x_values': series.xvalues,
                'y_values': series.yvalues
            }
            config['series'].append(series_config)
        
        return config
    
    def _extract_formatting_config(self, workbook: Workbook) -> Dict[str, Any]:
        """Extract formatting configuration from workbook"""
        return {
            'default_styles': {
                'header_style': {
                    'bold': True,
                    'size': 12
                },
                'data_style': {
                    'size': 11
                }
            }
        }
    
    def _extract_metadata_config(self, workbook: Workbook) -> Dict[str, Any]:
        """Extract metadata configuration from workbook"""
        metadata = {}
        
        if workbook.properties:
            metadata['properties'] = {
                'title': workbook.properties.title,
                'subject': workbook.properties.subject,
                'creator': workbook.properties.creator,
                'created': workbook.properties.created
            }
        
        return metadata
    
    def save_generation_history(self, file_path: str):
        """Save generation history to file"""
        with open(file_path, 'w') as f:
            json.dump(self.generation_history, f, indent=2)
        
        logger.info(f"Generation history saved to: {file_path}")
    
    def get_generation_statistics(self) -> Dict[str, Any]:
        """Get statistics about generation history"""
        if not self.generation_history:
            return {}
        
        total_generations = len(self.generation_history)
        total_files = len(set(gen['output_path'] for gen in self.generation_history))
        
        return {
            'total_generations': total_generations,
            'total_files_generated': total_files,
            'average_generations_per_file': total_generations / total_files if total_files > 0 else 0
        }
