#!/usr/bin/env python3
"""
Simplified Background Processing System for Excel Learning

This module provides a lightweight background processing system that works without
heavy AI/ML dependencies like TensorFlow. It focuses on basic Excel file analysis
and provides a foundation that can be extended when full AI capabilities are available.
"""

import os
import sys
import time
import json
import logging
import threading
import queue
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional, Callable, Union
from dataclasses import dataclass, asdict
from enum import Enum
from pathlib import Path
import pandas as pd
import numpy as np
import uuid

logger = logging.getLogger(__name__)

class TaskStatus(Enum):
    """Status of background learning tasks"""
    PENDING = "pending"
    PROCESSING = "processing"
    COMPLETED = "completed"
    FAILED = "failed"
    CANCELLED = "cancelled"

class NotificationType(Enum):
    """Types of notifications"""
    TASK_STARTED = "task_started"
    TASK_COMPLETED = "task_completed"
    TASK_FAILED = "task_failed"
    PROGRESS_UPDATE = "progress_update"
    SYSTEM_READY = "system_ready"

@dataclass
class LearningTask:
    """Represents a background learning task"""
    task_id: str
    file_path: str
    task_type: str  # 'excel_analysis', 'formula_learning', 'chart_learning', 'full_pipeline'
    priority: int = 1  # 1=low, 5=high
    created_at: datetime = None
    started_at: datetime = None
    completed_at: datetime = None
    status: TaskStatus = TaskStatus.PENDING
    progress: float = 0.0
    result: Dict[str, Any] = None
    error: str = None
    metadata: Dict[str, Any] = None
    
    def __post_init__(self):
        if self.created_at is None:
            self.created_at = datetime.now()
        if self.metadata is None:
            self.metadata = {}

@dataclass
class Notification:
    """Represents a notification to the user"""
    notification_id: str
    task_id: str
    notification_type: NotificationType
    message: str
    timestamp: datetime
    data: Dict[str, Any] = None
    read: bool = False
    
    def __post_init__(self):
        if self.data is None:
            self.data = {}

class SimpleExcelAnalyzer:
    """Simplified Excel analyzer that works without heavy dependencies"""
    
    def __init__(self):
        self.logger = logging.getLogger(__name__)
    
    def analyze_excel_file(self, file_path: str) -> Dict[str, Any]:
        """Analyze Excel file and extract basic information"""
        try:
            # Try to use openpyxl if available
            try:
                from openpyxl import load_workbook
                return self._analyze_with_openpyxl(file_path)
            except ImportError:
                # Fallback to pandas for basic analysis
                return self._analyze_with_pandas(file_path)
                
        except Exception as e:
            self.logger.error(f"Error analyzing Excel file {file_path}: {e}")
            raise
    
    def _analyze_with_openpyxl(self, file_path: str) -> Dict[str, Any]:
        """Analyze Excel file using openpyxl"""
        workbook = load_workbook(file_path, data_only=True)
        
        analysis = {
            'file_path': file_path,
            'file_size': os.path.getsize(file_path),
            'sheets': [],
            'formulas': [],
            'charts': [],
            'summary': {}
        }
        
        total_cells = 0
        total_formulas = 0
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet_data = {
                'name': sheet_name,
                'max_row': sheet.max_row,
                'max_column': sheet.max_column,
                'cells_with_data': 0,
                'formulas': []
            }
            
            # Analyze cells
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        sheet_data['cells_with_data'] += 1
                        total_cells += 1
                        
                        # Check for formulas
                        if cell.data_type == 'f' and cell.value:
                            formula_info = {
                                'cell': cell.coordinate,
                                'formula': cell.value,
                                'sheet': sheet_name
                            }
                            sheet_data['formulas'].append(formula_info)
                            analysis['formulas'].append(formula_info)
                            total_formulas += 1
            
            analysis['sheets'].append(sheet_data)
        
        # Summary
        analysis['summary'] = {
            'total_sheets': len(workbook.sheetnames),
            'total_cells': total_cells,
            'total_formulas': total_formulas,
            'analysis_timestamp': datetime.now().isoformat()
        }
        
        return analysis
    
    def _analyze_with_pandas(self, file_path: str) -> Dict[str, Any]:
        """Analyze Excel file using pandas (fallback method)"""
        try:
            # Read all sheets
            excel_file = pd.ExcelFile(file_path)
            
            analysis = {
                'file_path': file_path,
                'file_size': os.path.getsize(file_path),
                'sheets': [],
                'formulas': [],
                'charts': [],
                'summary': {}
            }
            
            total_cells = 0
            
            for sheet_name in excel_file.sheet_names:
                try:
                    df = pd.read_excel(file_path, sheet_name=sheet_name)
                    sheet_data = {
                        'name': sheet_name,
                        'rows': len(df),
                        'columns': len(df.columns),
                        'cells_with_data': df.count().sum(),
                        'formulas': []
                    }
                    
                    total_cells += sheet_data['cells_with_data']
                    analysis['sheets'].append(sheet_data)
                    
                except Exception as e:
                    logger.warning(f"Could not read sheet {sheet_name}: {e}")
            
            # Summary
            analysis['summary'] = {
                'total_sheets': len(excel_file.sheet_names),
                'total_cells': total_cells,
                'total_formulas': 0,  # Pandas doesn't preserve formulas
                'analysis_timestamp': datetime.now().isoformat(),
                'method': 'pandas_fallback'
            }
            
            return analysis
            
        except Exception as e:
            logger.error(f"Error analyzing with pandas: {e}")
            raise

class SimpleBackgroundProcessor:
    """
    Simplified background processor for Excel learning
    
    Features:
    - Queue-based task processing
    - Progress tracking and notifications
    - Multi-threaded processing
    - Task prioritization
    - Persistent storage of results
    - Works without heavy AI/ML dependencies
    """
    
    def __init__(self, 
                 max_workers: int = 2,
                 storage_path: str = "excel_learning_data",
                 enable_notifications: bool = True):
        """
        Initialize the background processor
        
        Args:
            max_workers: Maximum number of worker threads
            storage_path: Path to store persistent data
            enable_notifications: Whether to enable notifications
        """
        self.max_workers = max_workers
        self.storage_path = Path(storage_path)
        self.enable_notifications = enable_notifications
        
        # Create storage directory
        self.storage_path.mkdir(exist_ok=True)
        
        # Initialize components
        self.task_queue = queue.PriorityQueue()
        self.tasks: Dict[str, LearningTask] = {}
        self.notifications: List[Notification] = []
        self.notification_handlers: List[Callable[[Notification], None]] = []
        
        # Processing control
        self.running = False
        self.workers: List[threading.Thread] = []
        
        # Initialize analyzer
        self.analyzer = SimpleExcelAnalyzer()
        
        # Load persistent data
        self._load_persistent_data()
        
        logger.info(f"SimpleBackgroundProcessor initialized with {max_workers} workers")
    
    def _load_persistent_data(self):
        """Load persistent data from storage"""
        try:
            # Load tasks
            tasks_file = self.storage_path / "tasks.json"
            if tasks_file.exists():
                with open(tasks_file, 'r') as f:
                    tasks_data = json.load(f)
                    for task_dict in tasks_data:
                        task = self._dict_to_task(task_dict)
                        self.tasks[task.task_id] = task
            
            # Load notifications
            notifications_file = self.storage_path / "notifications.json"
            if notifications_file.exists():
                with open(notifications_file, 'r') as f:
                    notifications_data = json.load(f)
                    for notif_dict in notifications_data:
                        notification = self._dict_to_notification(notif_dict)
                        self.notifications.append(notification)
            
            logger.info(f"Loaded {len(self.tasks)} tasks and {len(self.notifications)} notifications")
            
        except Exception as e:
            logger.error(f"Error loading persistent data: {e}")
    
    def _save_persistent_data(self):
        """Save persistent data to storage"""
        try:
            # Save tasks
            tasks_file = self.storage_path / "tasks.json"
            tasks_data = [asdict(task) for task in self.tasks.values()]
            with open(tasks_file, 'w') as f:
                json.dump(tasks_data, f, indent=2, default=str)
            
            # Save notifications
            notifications_file = self.storage_path / "notifications.json"
            notifications_data = [asdict(notification) for notification in self.notifications]
            with open(notifications_file, 'w') as f:
                json.dump(notifications_data, f, indent=2, default=str)
            
            logger.debug("Persistent data saved successfully")
            
        except Exception as e:
            logger.error(f"Error saving persistent data: {e}")
    
    def add_notification_handler(self, handler: Callable[[Notification], None]):
        """Add a notification handler"""
        self.notification_handlers.append(handler)
    
    def _send_notification(self, notification: Notification):
        """Send notification to all handlers"""
        if not self.enable_notifications:
            return
        
        # Add to notifications list
        self.notifications.append(notification)
        
        # Send to handlers
        for handler in self.notification_handlers:
            try:
                handler(notification)
            except Exception as e:
                logger.error(f"Error in notification handler: {e}")
        
        # Save persistent data
        self._save_persistent_data()
    
    def submit_learning_task(self, 
                           file_path: str, 
                           task_type: str = "excel_analysis",
                           priority: int = 3,
                           metadata: Dict[str, Any] = None) -> str:
        """
        Submit a new learning task
        
        Args:
            file_path: Path to the Excel file
            task_type: Type of learning task
            priority: Task priority (1-5)
            metadata: Additional metadata
            
        Returns:
            Task ID
        """
        # Generate task ID
        task_id = str(uuid.uuid4())
        
        # Create task
        task = LearningTask(
            task_id=task_id,
            file_path=file_path,
            task_type=task_type,
            priority=priority,
            metadata=metadata or {}
        )
        
        # Add to tasks
        self.tasks[task_id] = task
        
        # Add to queue (priority queue uses negative priority for highest first)
        self.task_queue.put((-priority, task_id))
        
        # Send notification
        notification = Notification(
            notification_id=str(uuid.uuid4()),
            task_id=task_id,
            notification_type=NotificationType.TASK_STARTED,
            message=f"Task {task_id} submitted for {os.path.basename(file_path)}",
            timestamp=datetime.now()
        )
        self._send_notification(notification)
        
        logger.info(f"Task {task_id} submitted: {file_path}")
        return task_id
    
    def start_processing(self):
        """Start the background processing"""
        if self.running:
            logger.warning("Background processor already running")
            return
        
        self.running = True
        
        # Start worker threads
        for i in range(self.max_workers):
            worker = threading.Thread(target=self._worker_loop, name=f"Worker-{i}")
            worker.daemon = True
            worker.start()
            self.workers.append(worker)
        
        logger.info(f"Background processing started with {self.max_workers} workers")
        
        # Send system ready notification
        notification = Notification(
            notification_id=str(uuid.uuid4()),
            task_id="",
            notification_type=NotificationType.SYSTEM_READY,
            message="Background processor is ready",
            timestamp=datetime.now()
        )
        self._send_notification(notification)
    
    def stop_processing(self):
        """Stop the background processing"""
        if not self.running:
            return
        
        self.running = False
        
        # Wait for workers to finish
        for worker in self.workers:
            worker.join(timeout=5)
        
        # Save persistent data
        self._save_persistent_data()
        
        logger.info("Background processing stopped")
    
    def _worker_loop(self):
        """Main worker loop"""
        while self.running:
            try:
                # Get task from queue (with timeout)
                try:
                    priority, task_id = self.task_queue.get(timeout=1)
                except queue.Empty:
                    continue
                
                # Get task
                task = self.tasks.get(task_id)
                if not task:
                    continue
                
                # Process task
                self._process_task(task)
                
            except Exception as e:
                logger.error(f"Error in worker loop: {e}")
    
    def _process_task(self, task: LearningTask):
        """Process a single task"""
        try:
            # Update task status
            task.status = TaskStatus.PROCESSING
            task.started_at = datetime.now()
            task.progress = 0.0
            
            # Send notification
            notification = Notification(
                notification_id=str(uuid.uuid4()),
                task_id=task.task_id,
                notification_type=NotificationType.TASK_STARTED,
                message=f"Processing {os.path.basename(task.file_path)}",
                timestamp=datetime.now()
            )
            self._send_notification(notification)
            
            # Process based on task type
            if task.task_type == "excel_analysis":
                result = self._process_excel_analysis(task)
            elif task.task_type == "formula_learning":
                result = self._process_formula_learning(task)
            elif task.task_type == "chart_learning":
                result = self._process_chart_learning(task)
            elif task.task_type == "full_pipeline":
                result = self._process_full_pipeline(task)
            else:
                raise ValueError(f"Unknown task type: {task.task_type}")
            
            # Update task
            task.status = TaskStatus.COMPLETED
            task.completed_at = datetime.now()
            task.progress = 100.0
            task.result = result
            
            # Send completion notification
            notification = Notification(
                notification_id=str(uuid.uuid4()),
                task_id=task.task_id,
                notification_type=NotificationType.TASK_COMPLETED,
                message=f"Task completed: {os.path.basename(task.file_path)}",
                timestamp=datetime.now(),
                data={'result_summary': self._summarize_result(result)}
            )
            self._send_notification(notification)
            
            logger.info(f"Task {task.task_id} completed successfully")
            
        except Exception as e:
            # Handle failure
            task.status = TaskStatus.FAILED
            task.completed_at = datetime.now()
            task.error = str(e)
            
            # Send failure notification
            notification = Notification(
                notification_id=str(uuid.uuid4()),
                task_id=task.task_id,
                notification_type=NotificationType.TASK_FAILED,
                message=f"Task failed: {os.path.basename(task.file_path)} - {str(e)}",
                timestamp=datetime.now()
            )
            self._send_notification(notification)
            
            logger.error(f"Task {task.task_id} failed: {e}")
    
    def _process_excel_analysis(self, task: LearningTask) -> Dict[str, Any]:
        """Process Excel analysis task"""
        task.progress = 25.0
        
        # Analyze Excel file
        analysis = self.analyzer.analyze_excel_file(task.file_path)
        
        task.progress = 75.0
        
        # Add task metadata
        analysis['task_metadata'] = {
            'task_id': task.task_id,
            'task_type': task.task_type,
            'processing_time': datetime.now().isoformat()
        }
        
        task.progress = 100.0
        
        return {
            'excel_analysis': analysis,
            'processing_summary': {
                'sheets_analyzed': len(analysis['sheets']),
                'formulas_found': len(analysis['formulas']),
                'total_cells': analysis['summary']['total_cells']
            }
        }
    
    def _process_formula_learning(self, task: LearningTask) -> Dict[str, Any]:
        """Process formula learning task (simplified)"""
        task.progress = 25.0
        
        # Analyze Excel file first
        analysis = self.analyzer.analyze_excel_file(task.file_path)
        
        task.progress = 50.0
        
        # Extract formula patterns (simplified)
        learned_formulas = []
        for formula_info in analysis['formulas']:
            learned_formula = {
                'formula': formula_info['formula'],
                'cell': formula_info['cell'],
                'sheet': formula_info['sheet'],
                'pattern_type': 'basic',  # Simplified pattern recognition
                'confidence': 0.8
            }
            learned_formulas.append(learned_formula)
        
        task.progress = 100.0
        
        return {
            'formula_learning': {
                'learned_formulas': learned_formulas,
                'total_formulas': len(learned_formulas),
                'analysis_method': 'simplified_pattern_recognition'
            }
        }
    
    def _process_chart_learning(self, task: LearningTask) -> Dict[str, Any]:
        """Process chart learning task (simplified)"""
        task.progress = 25.0
        
        # For now, return basic chart detection
        # In a full implementation, this would analyze chart objects
        
        task.progress = 100.0
        
        return {
            'chart_learning': {
                'learned_charts': [],
                'chart_detection_method': 'simplified',
                'note': 'Full chart learning requires additional dependencies'
            }
        }
    
    def _process_full_pipeline(self, task: LearningTask) -> Dict[str, Any]:
        """Process full pipeline task"""
        task.progress = 20.0
        
        # Excel analysis
        excel_result = self._process_excel_analysis(task)
        
        task.progress = 50.0
        
        # Formula learning
        formula_result = self._process_formula_learning(task)
        
        task.progress = 80.0
        
        # Chart learning
        chart_result = self._process_chart_learning(task)
        
        task.progress = 100.0
        
        return {
            'excel_analysis': excel_result['excel_analysis'],
            'formula_learning': formula_result['formula_learning'],
            'chart_learning': chart_result['chart_learning'],
            'pipeline_summary': {
                'total_components': 3,
                'processing_complete': True
            }
        }
    
    def _summarize_result(self, result: Dict[str, Any]) -> Dict[str, Any]:
        """Create a summary of the result"""
        summary = {}
        
        if 'excel_analysis' in result:
            analysis = result['excel_analysis']
            summary['excel_analysis'] = {
                'sheets_analyzed': len(analysis.get('sheets', [])),
                'total_cells': analysis.get('summary', {}).get('total_cells', 0),
                'formulas_found': len(analysis.get('formulas', []))
            }
        
        if 'formula_learning' in result:
            formulas = result['formula_learning']
            summary['formula_learning'] = {
                'formulas_learned': len(formulas.get('learned_formulas', []))
            }
        
        if 'chart_learning' in result:
            charts = result['chart_learning']
            summary['chart_learning'] = {
                'charts_learned': len(charts.get('learned_charts', []))
            }
        
        return summary
    
    def get_task_status(self, task_id: str) -> Optional[LearningTask]:
        """Get the status of a specific task"""
        return self.tasks.get(task_id)
    
    def get_all_tasks(self) -> Dict[str, LearningTask]:
        """Get all tasks"""
        return self.tasks.copy()
    
    def get_notifications(self, unread_only: bool = False) -> List[Notification]:
        """Get notifications"""
        if unread_only:
            return [n for n in self.notifications if not n.read]
        return self.notifications.copy()
    
    def mark_notification_read(self, notification_id: str):
        """Mark a notification as read"""
        for notification in self.notifications:
            if notification.notification_id == notification_id:
                notification.read = True
                break
    
    def get_statistics(self) -> Dict[str, Any]:
        """Get processing statistics"""
        total_tasks = len(self.tasks)
        completed_tasks = len([t for t in self.tasks.values() if t.status == TaskStatus.COMPLETED])
        failed_tasks = len([t for t in self.tasks.values() if t.status == TaskStatus.FAILED])
        pending_tasks = len([t for t in self.tasks.values() if t.status == TaskStatus.PENDING])
        
        return {
            'total_tasks': total_tasks,
            'completed_tasks': completed_tasks,
            'failed_tasks': failed_tasks,
            'pending_tasks': pending_tasks,
            'success_rate': completed_tasks / total_tasks if total_tasks > 0 else 0,
            'active_workers': len([w for w in self.workers if w.is_alive()])
        }
    
    def _dict_to_task(self, task_dict: Dict[str, Any]) -> LearningTask:
        """Convert dictionary to LearningTask"""
        # Convert string status back to enum
        if 'status' in task_dict and isinstance(task_dict['status'], str):
            task_dict['status'] = TaskStatus(task_dict['status'])
        
        # Convert datetime strings back to datetime objects
        for field in ['created_at', 'started_at', 'completed_at']:
            if field in task_dict and task_dict[field]:
                if isinstance(task_dict[field], str):
                    task_dict[field] = datetime.fromisoformat(task_dict[field])
        
        return LearningTask(**task_dict)
    
    def _dict_to_notification(self, notif_dict: Dict[str, Any]) -> Notification:
        """Convert dictionary to Notification"""
        # Convert string notification type back to enum
        if 'notification_type' in notif_dict and isinstance(notif_dict['notification_type'], str):
            notif_dict['notification_type'] = NotificationType(notif_dict['notification_type'])
        
        # Convert datetime string back to datetime object
        if 'timestamp' in notif_dict and notif_dict['timestamp']:
            if isinstance(notif_dict['timestamp'], str):
                notif_dict['timestamp'] = datetime.fromisoformat(notif_dict['timestamp'])
        
        return Notification(**notif_dict)
    
    def cleanup_old_data(self, days_to_keep: int = 30):
        """Clean up old data"""
        cutoff_date = datetime.now() - timedelta(days=days_to_keep)
        
        # Clean up old tasks
        old_tasks = [task_id for task_id, task in self.tasks.items() 
                    if task.created_at < cutoff_date]
        for task_id in old_tasks:
            del self.tasks[task_id]
        
        # Clean up old notifications
        self.notifications = [n for n in self.notifications 
                            if n.timestamp > cutoff_date]
        
        # Save persistent data
        self._save_persistent_data()
        
        logger.info(f"Cleaned up {len(old_tasks)} old tasks and old notifications")
    
    def export_learning_results(self, output_path: str):
        """Export learning results to file"""
        try:
            export_data = {
                'export_timestamp': datetime.now().isoformat(),
                'tasks': [asdict(task) for task in self.tasks.values()],
                'statistics': self.get_statistics()
            }
            
            with open(output_path, 'w') as f:
                json.dump(export_data, f, indent=2, default=str)
            
            logger.info(f"Learning results exported to {output_path}")
            
        except Exception as e:
            logger.error(f"Error exporting learning results: {e}")
            raise
