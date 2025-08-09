"""
Excel Analyzer for AI Learning

This module analyzes Excel files to extract patterns, structure, and features
that can be learned by AI models to generate similar files.
"""

import pandas as pd
import numpy as np
import openpyxl
from openpyxl.chart import Chart
from openpyxl.styles import Font, PatternFill, Border, Alignment
from openpyxl.utils import get_column_letter
import json
import logging
from typing import Dict, List, Tuple, Any, Optional
from dataclasses import dataclass
from pathlib import Path
import cv2
from PIL import Image
import io
from .formula_learner import FormulaLearner, FormulaLogic, FilterCondition

logger = logging.getLogger(__name__)

@dataclass
class ExcelStructure:
    """Represents the structure of an Excel file"""
    sheet_names: List[str]
    data_ranges: Dict[str, Dict[str, Any]]
    charts: Dict[str, List[Dict[str, Any]]]
    formatting: Dict[str, Dict[str, Any]]
    formulas: Dict[str, List[Dict[str, Any]]]
    metadata: Dict[str, Any]

@dataclass
class DataPattern:
    """Represents a data pattern found in Excel"""
    pattern_type: str  # 'sequential', 'random', 'formula', 'categorical'
    data_type: str     # 'numeric', 'text', 'date', 'boolean'
    range_info: Dict[str, Any]
    statistics: Dict[str, float]
    relationships: List[Dict[str, Any]]

class ExcelAnalyzer:
    """
    Analyzes Excel files to extract patterns and structure for AI learning
    """
    
    def __init__(self):
        self.supported_chart_types = [
            'scatter', 'line', 'bar', 'column', 'pie', 'area', 'doughnut'
        ]
        self.formula_learner = FormulaLearner()
        
    def analyze_excel_file(self, file_path: str) -> ExcelStructure:
        """
        Comprehensive analysis of an Excel file
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            ExcelStructure containing all extracted information
        """
        logger.info(f"Analyzing Excel file: {file_path}")
        
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            structure = ExcelStructure(
                sheet_names=workbook.sheetnames,
                data_ranges={},
                charts={},
                formatting={},
                formulas={},
                metadata={}
            )
            
            # Analyze each sheet
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                logger.info(f"Analyzing sheet: {sheet_name}")
                
                # Extract data ranges and patterns
                structure.data_ranges[sheet_name] = self._analyze_data_ranges(sheet)
                
                # Extract charts
                structure.charts[sheet_name] = self._analyze_charts(sheet)
                
                # Extract formatting
                structure.formatting[sheet_name] = self._analyze_formatting(sheet)
                
                # Extract formulas
                structure.formulas[sheet_name] = self._analyze_formulas(sheet)
            
            # Extract metadata
            structure.metadata = self._extract_metadata(workbook)
            
            return structure
            
        except Exception as e:
            logger.error(f"Error analyzing Excel file {file_path}: {str(e)}")
            raise
    
    def _analyze_data_ranges(self, sheet) -> Dict[str, Any]:
        """Analyze data ranges and patterns in a sheet"""
        data_info = {
            'ranges': [],
            'patterns': [],
            'statistics': {}
        }
        
        # Find data ranges
        data_ranges = self._find_data_ranges(sheet)
        
        for range_info in data_ranges:
            # Extract data from range
            data = self._extract_range_data(sheet, range_info)
            
            # Analyze patterns
            patterns = self._analyze_data_patterns(data)
            
            # Calculate statistics
            stats = self._calculate_statistics(data)
            
            data_info['ranges'].append({
                'range': range_info,
                'data': data,
                'patterns': patterns,
                'statistics': stats
            })
        
        return data_info
    
    def _find_data_ranges(self, sheet) -> List[Dict[str, Any]]:
        """Find all data ranges in the sheet"""
        ranges = []
        
        # Find the used range
        min_row = sheet.min_row
        max_row = sheet.max_row
        min_col = sheet.min_column
        max_col = sheet.max_column
        
        if min_row is None or max_row is None:
            return ranges
        
        # Look for continuous data blocks
        current_range = None
        
        for row in range(min_row, max_row + 1):
            row_has_data = False
            for col in range(min_col, max_col + 1):
                cell = sheet.cell(row=row, column=col)
                if cell.value is not None:
                    row_has_data = True
                    break
            
            if row_has_data:
                if current_range is None:
                    current_range = {
                        'start_row': row,
                        'end_row': row,
                        'start_col': min_col,
                        'end_col': max_col
                    }
                else:
                    current_range['end_row'] = row
            else:
                if current_range is not None:
                    ranges.append(current_range)
                    current_range = None
        
        if current_range is not None:
            ranges.append(current_range)
        
        return ranges
    
    def _extract_range_data(self, sheet, range_info: Dict[str, Any]) -> pd.DataFrame:
        """Extract data from a specific range"""
        data = []
        headers = []
        
        for row in range(range_info['start_row'], range_info['end_row'] + 1):
            row_data = []
            for col in range(range_info['start_col'], range_info['end_col'] + 1):
                cell = sheet.cell(row=row, column=col)
                row_data.append(cell.value)
            
            if row == range_info['start_row']:
                headers = row_data
            else:
                data.append(row_data)
        
        return pd.DataFrame(data, columns=headers)
    
    def _analyze_data_patterns(self, data: pd.DataFrame) -> List[DataPattern]:
        """Analyze patterns in the data"""
        patterns = []
        
        for column in data.columns:
            col_data = data[column].dropna()
            
            if len(col_data) == 0:
                continue
            
            # Determine data type
            data_type = self._determine_data_type(col_data)
            
            # Analyze pattern type
            pattern_type = self._determine_pattern_type(col_data)
            
            # Find relationships with other columns
            relationships = self._find_relationships(data, column)
            
            pattern = DataPattern(
                pattern_type=pattern_type,
                data_type=data_type,
                range_info={'column': column},
                statistics=self._calculate_column_statistics(col_data),
                relationships=relationships
            )
            
            patterns.append(pattern)
        
        return patterns
    
    def _determine_data_type(self, data: pd.Series) -> str:
        """Determine the data type of a column"""
        if data.dtype in ['int64', 'float64']:
            return 'numeric'
        elif data.dtype == 'bool':
            return 'boolean'
        elif data.dtype == 'datetime64[ns]':
            return 'date'
        else:
            return 'text'
    
    def _determine_pattern_type(self, data: pd.Series) -> str:
        """Determine the pattern type of a column"""
        if len(data) < 2:
            return 'constant'
        
        # Check if sequential
        if data.dtype in ['int64', 'float64']:
            diff = data.diff().dropna()
            if len(diff) > 0 and abs(diff.std()) < 0.1:
                return 'sequential'
            
            # Check for formula patterns
            if self._is_formula_pattern(data):
                return 'formula'
        
        # Check for categorical
        unique_ratio = len(data.unique()) / len(data)
        if unique_ratio < 0.3:
            return 'categorical'
        
        return 'random'
    
    def _is_formula_pattern(self, data: pd.Series) -> bool:
        """Check if data follows a formula pattern"""
        if len(data) < 3:
            return False
        
        # Check for linear patterns
        x = np.arange(len(data))
        slope, intercept = np.polyfit(x, data, 1)
        r_squared = 1 - np.sum((data - (slope * x + intercept)) ** 2) / np.sum((data - data.mean()) ** 2)
        
        return r_squared > 0.95
    
    def _find_relationships(self, data: pd.DataFrame, column: str) -> List[Dict[str, Any]]:
        """Find relationships between columns"""
        relationships = []
        
        for other_col in data.columns:
            if other_col == column:
                continue
            
            # Calculate correlation for numeric columns
            if data[column].dtype in ['int64', 'float64'] and data[other_col].dtype in ['int64', 'float64']:
                correlation = data[column].corr(data[other_col])
                if abs(correlation) > 0.7:
                    relationships.append({
                        'column': other_col,
                        'type': 'correlation',
                        'strength': correlation
                    })
        
        return relationships
    
    def _calculate_statistics(self, data: pd.DataFrame) -> Dict[str, float]:
        """Calculate overall statistics for the data"""
        stats = {
            'total_rows': len(data),
            'total_columns': len(data.columns),
            'missing_values': data.isnull().sum().sum(),
            'missing_percentage': (data.isnull().sum().sum() / (len(data) * len(data.columns))) * 100
        }
        
        # Numeric statistics
        numeric_cols = data.select_dtypes(include=[np.number]).columns
        if len(numeric_cols) > 0:
            stats['numeric_columns'] = len(numeric_cols)
            stats['numeric_mean'] = data[numeric_cols].mean().mean()
            stats['numeric_std'] = data[numeric_cols].std().mean()
        
        return stats
    
    def _calculate_column_statistics(self, data: pd.Series) -> Dict[str, float]:
        """Calculate statistics for a single column"""
        stats = {
            'count': len(data),
            'unique_count': len(data.unique()),
            'missing_count': data.isnull().sum()
        }
        
        if data.dtype in ['int64', 'float64']:
            stats.update({
                'mean': data.mean(),
                'std': data.std(),
                'min': data.min(),
                'max': data.max(),
                'median': data.median()
            })
        
        return stats
    
    def _analyze_charts(self, sheet) -> List[Dict[str, Any]]:
        """Analyze charts in the sheet"""
        charts = []
        
        for chart in sheet._charts:
            chart_info = {
                'type': type(chart).__name__,
                'title': getattr(chart, 'title', None),
                'x_axis': self._extract_axis_info(chart, 'x'),
                'y_axis': self._extract_axis_info(chart, 'y'),
                'series': self._extract_series_info(chart),
                'position': self._extract_chart_position(chart)
            }
            charts.append(chart_info)
        
        return charts
    
    def _extract_axis_info(self, chart, axis_type: str) -> Dict[str, Any]:
        """Extract axis information from chart"""
        axis = getattr(chart, f'{axis_type}_axis', None)
        if axis is None:
            return {}
        
        return {
            'title': getattr(axis, 'title', None),
            'min': getattr(axis, 'min', None),
            'max': getattr(axis, 'max', None),
            'major_unit': getattr(axis, 'major_unit', None),
            'minor_unit': getattr(axis, 'minor_unit', None)
        }
    
    def _extract_series_info(self, chart) -> List[Dict[str, Any]]:
        """Extract series information from chart"""
        series_info = []
        
        for series in chart.series:
            info = {
                'name': getattr(series, 'title', None),
                'x_values': getattr(series, 'xvalues', None),
                'y_values': getattr(series, 'yvalues', None),
                'marker': getattr(series, 'marker', None)
            }
            series_info.append(info)
        
        return series_info
    
    def _extract_chart_position(self, chart) -> Dict[str, Any]:
        """Extract chart position information"""
        anchor = getattr(chart, '_anchor', None)
        if anchor is None:
            return {}
        
        return {
            'left': getattr(anchor, 'left', None),
            'top': getattr(anchor, 'top', None),
            'width': getattr(anchor, 'width', None),
            'height': getattr(anchor, 'height', None)
        }
    
    def _analyze_formatting(self, sheet) -> Dict[str, Any]:
        """Analyze cell formatting in the sheet"""
        formatting = {
            'fonts': {},
            'fills': {},
            'borders': {},
            'alignments': {}
        }
        
        # Analyze formatting for used range
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    # Font analysis
                    if cell.font:
                        font_key = f"{cell.font.name}_{cell.font.size}_{cell.font.bold}_{cell.font.italic}"
                        formatting['fonts'][font_key] = formatting['fonts'].get(font_key, 0) + 1
                    
                    # Fill analysis
                    if cell.fill:
                        fill_key = str(cell.fill.start_color.rgb)
                        formatting['fills'][fill_key] = formatting['fills'].get(fill_key, 0) + 1
                    
                    # Border analysis
                    if cell.border:
                        border_key = str(cell.border.style)
                        formatting['borders'][border_key] = formatting['borders'].get(border_key, 0) + 1
                    
                    # Alignment analysis
                    if cell.alignment:
                        align_key = f"{cell.alignment.horizontal}_{cell.alignment.vertical}"
                        formatting['alignments'][align_key] = formatting['alignments'].get(align_key, 0) + 1
        
        return formatting
    
    def _analyze_formulas(self, sheet) -> List[Dict[str, Any]]:
        """Analyze formulas in the sheet with advanced logic learning"""
        formulas = []
        
        # Extract sheet data for context
        sheet_data = []
        for row in sheet.iter_rows():
            row_data = []
            for cell in row:
                row_data.append(cell.value)
            sheet_data.append(row_data)
        
        # Convert to DataFrame for context analysis
        df = pd.DataFrame(sheet_data)
        
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None and isinstance(cell.value, str) and cell.value.startswith('='):
                    # Basic formula info
                    formula_info = {
                        'cell': cell.coordinate,
                        'formula': cell.value,
                        'result': cell.value if not cell.data_type == 'f' else None
                    }
                    
                    # Learn formula logic using FormulaLearner
                    try:
                        formula_logic = self.formula_learner.learn_formula_logic(cell.value, df)
                        formula_info['learned_logic'] = {
                            'formula_type': formula_logic.formula_type.value,
                            'operation': formula_logic.operation,
                            'source_range': formula_logic.source_range,
                            'target_range': formula_logic.target_range,
                            'conditions': formula_logic.conditions,
                            'parameters': formula_logic.parameters,
                            'dependencies': formula_logic.dependencies,
                            'result_type': formula_logic.result_type,
                            'confidence': formula_logic.confidence
                        }
                    except Exception as e:
                        logger.warning(f"Could not learn logic for formula {cell.value}: {e}")
                        formula_info['learned_logic'] = None
                    
                    formulas.append(formula_info)
        
        return formulas
    
    def _extract_metadata(self, workbook) -> Dict[str, Any]:
        """Extract metadata from workbook"""
        metadata = {
            'properties': {},
            'defined_names': [],
            'external_links': []
        }
        
        # Extract properties
        if workbook.properties:
            props = workbook.properties
            metadata['properties'] = {
                'title': getattr(props, 'title', None),
                'subject': getattr(props, 'subject', None),
                'creator': getattr(props, 'creator', None),
                'created': getattr(props, 'created', None),
                'modified': getattr(props, 'modified', None)
            }
        
        # Extract defined names
        for name in workbook.defined_names.definedName:
            metadata['defined_names'].append({
                'name': name.name,
                'value': name.attr_text
            })
        
        return metadata
    
    def save_analysis(self, structure: ExcelStructure, output_path: str):
        """Save analysis results to JSON file"""
        # Convert to serializable format
        analysis_data = {
            'sheet_names': structure.sheet_names,
            'data_ranges': self._serialize_data_ranges(structure.data_ranges),
            'charts': structure.charts,
            'formatting': structure.formatting,
            'formulas': structure.formulas,
            'metadata': structure.metadata
        }
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(analysis_data, f, indent=2, default=str)
        
        logger.info(f"Analysis saved to: {output_path}")
    
    def _serialize_data_ranges(self, data_ranges: Dict[str, Any]) -> Dict[str, Any]:
        """Convert data ranges to serializable format"""
        serialized = {}
        
        for sheet_name, sheet_data in data_ranges.items():
            serialized[sheet_name] = {
                'ranges': [],
                'patterns': [],
                'statistics': sheet_data['statistics']
            }
            
            for range_data in sheet_data['ranges']:
                serialized_range = {
                    'range': range_data['range'],
                    'statistics': range_data['statistics'],
                    'patterns': []
                }
                
                for pattern in range_data['patterns']:
                    serialized_pattern = {
                        'pattern_type': pattern.pattern_type,
                        'data_type': pattern.data_type,
                        'range_info': pattern.range_info,
                        'statistics': pattern.statistics,
                        'relationships': pattern.relationships
                    }
                    serialized_range['patterns'].append(serialized_pattern)
                
                serialized[sheet_name]['ranges'].append(serialized_range)
        
        return serialized
    
    def learn_filtering_operations(self, data: pd.DataFrame, filter_conditions: List[Dict[str, Any]]) -> List[FilterCondition]:
        """
        Learn filtering operations from data and conditions
        
        Args:
            data: DataFrame being filtered
            filter_conditions: List of filter conditions applied
            
        Returns:
            List of learned FilterCondition objects
        """
        return self.formula_learner.learn_filtering_operations(data, filter_conditions)
    
    def apply_learned_filters(self, data: pd.DataFrame, filters: List[FilterCondition] = None) -> pd.DataFrame:
        """
        Apply learned filtering operations to data
        
        Args:
            data: DataFrame to filter
            filters: List of FilterCondition objects (uses learned filters if None)
            
        Returns:
            Filtered DataFrame
        """
        return self.formula_learner.apply_learned_filters(data, filters)
    
    def get_learned_formula_patterns(self) -> Dict[str, Any]:
        """Get all learned formula patterns"""
        return self.formula_learner.get_learned_patterns()
    
    def get_learned_filter_patterns(self) -> List[FilterCondition]:
        """Get all learned filter patterns"""
        return self.formula_learner.get_filter_patterns()
    
    def save_learned_patterns(self, filepath: str):
        """Save learned patterns to file"""
        self.formula_learner.save_learned_patterns(filepath)
    
    def load_learned_patterns(self, filepath: str):
        """Load learned patterns from file"""
        self.formula_learner.load_learned_patterns(filepath)
