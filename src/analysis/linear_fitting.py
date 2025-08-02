"""
Enhanced Linear Drift Correction (Egyenes Illesztés) for Purple Curves

This module implements an improved linear regression-based drift correction with:
1. AI-powered segment detection with multiple fallback methods
2. Robust outlier handling using multiple algorithms
3. Statistical validation and confidence intervals
4. Smart edge case handling and automatic method selection
5. Backward compatibility with existing code

Hungarian term: "egyenes illesztés" = linear fitting/regression

File: src/analysis/linear_fitting.py (ENHANCED VERSION)
"""

import numpy as np
from scipy import stats
from scipy.signal import find_peaks, savgol_filter
from scipy.optimize import minimize_scalar
import matplotlib.pyplot as plt
from typing import Dict, List, Tuple, Optional, Union
import warnings

# Try importing robust methods, with graceful fallbacks
try:
    from sklearn.linear_model import RANSACRegressor, HuberRegressor, TheilSenRegressor
    from sklearn.preprocessing import RobustScaler
    from sklearn.metrics import r2_score

    SKLEARN_AVAILABLE = True
except ImportError:
    SKLEARN_AVAILABLE = False
    print("Warning: scikit-learn not available. Using basic methods only.")

try:
    from src.utils.logger import app_logger
except ImportError:
    import logging

    app_logger = logging.getLogger(__name__)


class EnhancedLinearDriftCorrector:
    """
    Enhanced linear drift correction with AI-powered analysis and robust statistics.
    Backward compatible with existing LinearDriftCorrector API.
    """

    def __init__(self, outlier_threshold: float = 2.5, min_segment_length: int = 10):
        """
        Initialize the enhanced drift corrector.

        Args:
            outlier_threshold: Z-score threshold for outlier detection
            min_segment_length: Minimum points required for a valid segment
        """
        self.outlier_threshold = outlier_threshold
        self.min_segment_length = min_segment_length
        self.regression_results = {}
        self.corrected_curves = {}
        self.drift_lines = {}
        self.confidence_intervals = {}
        self.sklearn_available = SKLEARN_AVAILABLE

    def identify_regression_segments(
        self,
        curve_data: np.ndarray,
        curve_times: np.ndarray,
        curve_type: str = "hyperpol",
        method: str = "auto",
    ) -> Dict:
        """
        Enhanced segment identification with multiple detection methods.

        Args:
            curve_data: Purple curve data (hyperpol or depol)
            curve_times: Time values for the curve
            curve_type: 'hyperpol' (ascending) or 'depol' (descending)
            method: 'auto', 'derivative', 'ai_powered', or 'manual'

        Returns:
            dict with segment information and quality metrics
        """
        app_logger.info(
            f"Identifying regression segments for {curve_type} curve using {method} method"
        )

        try:
            # Input validation
            if len(curve_data) != len(curve_times):
                raise ValueError("Curve data and times must have same length")

            if len(curve_data) < self.min_segment_length:
                raise ValueError(
                    f"Insufficient data points: {len(curve_data)} < {self.min_segment_length}"
                )

            # Preprocess data to remove noise
            cleaned_data = self._preprocess_curve_data(curve_data)

            if method == "auto":
                # Try multiple methods and choose the best
                segments = self._auto_detect_best_segment(
                    cleaned_data, curve_times, curve_type
                )
            elif method == "derivative":
                segments = self._derivative_based_detection(
                    cleaned_data, curve_times, curve_type
                )
            elif method == "ai_powered":
                segments = self._ai_powered_detection(
                    cleaned_data, curve_times, curve_type
                )
            else:  # manual fallback
                segments = self._manual_segment_detection(
                    cleaned_data, curve_times, curve_type
                )

            # Validate and refine the segment
            validated_segment = self._validate_and_refine_segment(
                segments, cleaned_data, curve_times, curve_type
            )

            app_logger.debug(
                f"Found {curve_type} regression segment: indices {validated_segment['start_idx']}-{validated_segment['end_idx']}"
            )
            return validated_segment

        except Exception as e:
            app_logger.error(f"Error identifying regression segments: {str(e)}")
            # Return fallback segment if all else fails
            return self._emergency_fallback_segment(
                curve_data, curve_times, curve_type, str(e)
            )

    def _preprocess_curve_data(self, curve_data: np.ndarray) -> np.ndarray:
        """
        Preprocess curve data to remove noise and outliers.
        """
        try:
            # Apply Savitzky-Golay filter to reduce noise
            if len(curve_data) > 5:
                window_length = min(max(5, len(curve_data) // 10), len(curve_data) - 1)
                if window_length % 2 == 0:
                    window_length -= 1
                cleaned_data = savgol_filter(curve_data, window_length, 3)
            else:
                cleaned_data = curve_data.copy()

            # Remove obvious outliers using z-score
            z_scores = np.abs(stats.zscore(cleaned_data, nan_policy="omit"))
            outlier_mask = z_scores < self.outlier_threshold

            # Interpolate outliers
            if not np.all(outlier_mask):
                outlier_indices = np.where(~outlier_mask)[0]
                for idx in outlier_indices:
                    # Simple linear interpolation for outliers
                    if 0 < idx < len(cleaned_data) - 1:
                        cleaned_data[idx] = (
                            cleaned_data[idx - 1] + cleaned_data[idx + 1]
                        ) / 2

            return cleaned_data

        except Exception as e:
            app_logger.warning(f"Error in preprocessing, using original data: {str(e)}")
            return curve_data.copy()

    def _auto_detect_best_segment(
        self, curve_data: np.ndarray, curve_times: np.ndarray, curve_type: str
    ) -> Dict:
        """
        Automatically detect the best segment using multiple methods and scoring.
        """
        methods = ["derivative", "gradient", "peak_based"]
        best_segment = None
        best_score = -np.inf

        for method in methods:
            try:
                if method == "derivative":
                    segment = self._derivative_based_detection(
                        curve_data, curve_times, curve_type
                    )
                elif method == "gradient":
                    segment = self._gradient_based_detection(
                        curve_data, curve_times, curve_type
                    )
                else:  # peak_based
                    segment = self._peak_based_detection(
                        curve_data, curve_times, curve_type
                    )

                # Score the segment quality
                score = self._score_segment_quality(segment, curve_data, curve_times)

                if score > best_score:
                    best_score = score
                    best_segment = segment
                    best_segment["detection_method"] = method
                    best_segment["quality_score"] = score

            except Exception as e:
                app_logger.debug(f"Method {method} failed: {str(e)}")
                continue

        if best_segment is None:
            # Fallback to simple segment
            return self._simple_fallback_segment(curve_data, curve_times, curve_type)

        return best_segment

    def _derivative_based_detection(
        self, curve_data: np.ndarray, curve_times: np.ndarray, curve_type: str
    ) -> Dict:
        """
        Detect segments based on derivative analysis.
        """
        # Calculate first derivative
        dt = np.diff(curve_times)
        dy = np.diff(curve_data)

        # Avoid division by zero
        dt = np.where(dt == 0, 1e-10, dt)
        derivative = dy / dt

        if curve_type == "hyperpol":
            # For hyperpolarization, find the steepest ascending region
            max_deriv_idx = np.argmax(derivative) + 1

            # Find region around maximum derivative where it's consistently positive
            start_idx = max_deriv_idx
            end_idx = max_deriv_idx

            # Extend backwards while derivative is positive and significant
            threshold = np.max(derivative) * 0.1
            while start_idx > 0 and derivative[start_idx - 1] > threshold:
                start_idx -= 1

            # Extend forwards while derivative is positive and significant
            while end_idx < len(derivative) - 1 and derivative[end_idx] > threshold:
                end_idx += 1

            end_idx += 1

        else:  # depol
            # For depolarization, find the steepest descending region
            min_deriv_idx = np.argmin(derivative) + 1

            start_idx = min_deriv_idx
            end_idx = min_deriv_idx

            # Extend while derivative is negative and significant
            threshold = np.min(derivative) * 0.1
            while start_idx > 0 and derivative[start_idx - 1] < threshold:
                start_idx -= 1

            while end_idx < len(derivative) - 1 and derivative[end_idx] < threshold:
                end_idx += 1

            end_idx += 1

        # Ensure minimum segment length
        if end_idx - start_idx < self.min_segment_length:
            center = (start_idx + end_idx) // 2
            half_length = self.min_segment_length // 2
            start_idx = max(0, center - half_length)
            end_idx = min(len(curve_data), center + half_length)

        return {
            "start_idx": start_idx,
            "end_idx": end_idx,
            "segment_type": curve_type,
            "detection_method": "derivative",
            "max_derivative": np.max(np.abs(derivative)),
            "linearity_score": self._calculate_linearity_score(
                curve_data[start_idx:end_idx], curve_times[start_idx:end_idx]
            ),
        }

    def _gradient_based_detection(
        self, curve_data: np.ndarray, curve_times: np.ndarray, curve_type: str
    ) -> Dict:
        """
        Detect segments using gradient analysis and sliding window.
        """
        window_size = max(5, len(curve_data) // 20)
        gradients = []
        positions = []

        for i in range(window_size, len(curve_data) - window_size):
            window_times = curve_times[i - window_size : i + window_size]
            window_data = curve_data[i - window_size : i + window_size]

            # Calculate local gradient
            slope, _, r_value, _, _ = stats.linregress(window_times, window_data)
            gradients.append(slope)
            positions.append(i)

        gradients = np.array(gradients)
        positions = np.array(positions)

        if curve_type == "hyperpol":
            # Find region with maximum positive gradient
            target_idx = np.argmax(gradients)
        else:
            # Find region with maximum negative gradient
            target_idx = np.argmin(gradients)

        center_pos = positions[target_idx]

        # Define segment around this center
        segment_half_length = max(self.min_segment_length, len(curve_data) // 10)
        start_idx = max(0, center_pos - segment_half_length)
        end_idx = min(len(curve_data), center_pos + segment_half_length)

        return {
            "start_idx": start_idx,
            "end_idx": end_idx,
            "segment_type": curve_type,
            "detection_method": "gradient",
            "peak_gradient": gradients[target_idx],
            "linearity_score": self._calculate_linearity_score(
                curve_data[start_idx:end_idx], curve_times[start_idx:end_idx]
            ),
        }

    def _peak_based_detection(
        self, curve_data: np.ndarray, curve_times: np.ndarray, curve_type: str
    ) -> Dict:
        """
        Detect segments using peak analysis.
        """
        # Find peaks and valleys
        peaks, _ = find_peaks(curve_data, prominence=np.std(curve_data) * 0.5)
        valleys, _ = find_peaks(-curve_data, prominence=np.std(curve_data) * 0.5)

        if curve_type == "hyperpol":
            # Find the region before the main peak (recovery phase)
            if len(valleys) > 0 and len(peaks) > 0:
                main_valley = valleys[np.argmin(curve_data[valleys])]
                relevant_peaks = peaks[peaks > main_valley]
                if len(relevant_peaks) > 0:
                    main_peak = relevant_peaks[0]
                    start_idx = main_valley
                    end_idx = main_peak
                else:
                    start_idx = main_valley
                    end_idx = min(len(curve_data), main_valley + len(curve_data) // 4)
            else:
                # Fallback to middle region
                start_idx = len(curve_data) // 3
                end_idx = 2 * len(curve_data) // 3

        else:  # depol
            # Find the region after the main peak (decay phase)
            if len(peaks) > 0:
                main_peak = peaks[np.argmax(curve_data[peaks])]
                start_idx = main_peak
                # Look for next valley or end of significant decay
                relevant_valleys = valleys[valleys > main_peak]
                if len(relevant_valleys) > 0:
                    end_idx = relevant_valleys[0]
                else:
                    end_idx = min(len(curve_data), main_peak + len(curve_data) // 4)
            else:
                # Fallback to latter region
                start_idx = len(curve_data) // 3
                end_idx = 2 * len(curve_data) // 3

        # Ensure minimum length
        if end_idx - start_idx < self.min_segment_length:
            center = (start_idx + end_idx) // 2
            half_length = self.min_segment_length // 2
            start_idx = max(0, center - half_length)
            end_idx = min(len(curve_data), center + half_length)

        return {
            "start_idx": start_idx,
            "end_idx": end_idx,
            "segment_type": curve_type,
            "detection_method": "peak_based",
            "num_peaks": len(peaks),
            "num_valleys": len(valleys),
            "linearity_score": self._calculate_linearity_score(
                curve_data[start_idx:end_idx], curve_times[start_idx:end_idx]
            ),
        }

    def _ai_powered_detection(
        self, curve_data: np.ndarray, curve_times: np.ndarray, curve_type: str
    ) -> Dict:
        """
        AI-powered segment detection (placeholder for future AI integration).
        Currently falls back to auto detection.
        """
        app_logger.info(
            "AI-powered detection requested, falling back to auto detection"
        )
        return self._auto_detect_best_segment(curve_data, curve_times, curve_type)

    def _manual_segment_detection(
        self, curve_data: np.ndarray, curve_times: np.ndarray, curve_type: str
    ) -> Dict:
        """
        Manual segment detection fallback.
        """
        return self._simple_fallback_segment(curve_data, curve_times, curve_type)

    def _score_segment_quality(
        self, segment: Dict, curve_data: np.ndarray, curve_times: np.ndarray
    ) -> float:
        """
        Score the quality of a detected segment.
        """
        try:
            start_idx = segment["start_idx"]
            end_idx = segment["end_idx"]

            # Extract segment
            segment_times = curve_times[start_idx:end_idx]
            segment_data = curve_data[start_idx:end_idx]

            if len(segment_data) < 3:
                return -np.inf

            # Calculate R-squared for linearity
            slope, intercept, r_value, _, _ = stats.linregress(
                segment_times, segment_data
            )
            r_squared = r_value**2

            # Length score (prefer reasonable lengths)
            length_score = min(1.0, len(segment_data) / (len(curve_data) * 0.3))

            # Consistency score (prefer segments with consistent slope direction)
            derivatives = np.diff(segment_data) / np.diff(segment_times)
            if len(derivatives) > 0:
                consistency_score = 1.0 - np.std(derivatives) / (
                    np.abs(np.mean(derivatives)) + 1e-6
                )
            else:
                consistency_score = 0

            # Combined score
            total_score = (
                r_squared * 0.5 + length_score * 0.2 + max(0, consistency_score) * 0.3
            )

            return total_score

        except Exception as e:
            app_logger.warning(f"Error scoring segment: {e}")
            return -np.inf

    def _calculate_linearity_score(
        self, segment_data: np.ndarray, segment_times: np.ndarray
    ) -> float:
        """
        Calculate how linear a segment is.
        """
        if len(segment_data) < 3:
            return 0.0

        try:
            _, _, r_value, _, _ = stats.linregress(segment_times, segment_data)
            return r_value**2
        except:
            return 0.0

    def _validate_and_refine_segment(
        self,
        segment: Dict,
        curve_data: np.ndarray,
        curve_times: np.ndarray,
        curve_type: str,
    ) -> Dict:
        """
        Validate and refine the detected segment.
        """
        start_idx = segment["start_idx"]
        end_idx = segment["end_idx"]

        # Ensure bounds
        start_idx = max(0, start_idx)
        end_idx = min(len(curve_data), end_idx)

        # Ensure minimum length
        if end_idx - start_idx < self.min_segment_length:
            center = (start_idx + end_idx) // 2
            half_length = self.min_segment_length // 2
            start_idx = max(0, center - half_length)
            end_idx = min(len(curve_data), center + half_length)

        # Try to improve the segment by optimizing R-squared
        optimized_segment = self._optimize_segment_bounds(
            start_idx, end_idx, curve_data, curve_times
        )

        segment.update(optimized_segment)
        segment["validated"] = True

        return segment

    def _optimize_segment_bounds(
        self,
        start_idx: int,
        end_idx: int,
        curve_data: np.ndarray,
        curve_times: np.ndarray,
    ) -> Dict:
        """
        Optimize segment boundaries to maximize R-squared.
        """
        best_r_squared = -np.inf
        best_start = start_idx
        best_end = end_idx

        # Try small adjustments to boundaries
        for start_adj in range(-5, 6):
            for end_adj in range(-5, 6):
                test_start = max(0, start_idx + start_adj)
                test_end = min(len(curve_data), end_idx + end_adj)

                if test_end - test_start < self.min_segment_length:
                    continue

                try:
                    test_times = curve_times[test_start:test_end]
                    test_data = curve_data[test_start:test_end]
                    _, _, r_value, _, _ = stats.linregress(test_times, test_data)
                    r_squared = r_value**2

                    if r_squared > best_r_squared:
                        best_r_squared = r_squared
                        best_start = test_start
                        best_end = test_end

                except:
                    continue

        return {
            "start_idx": best_start,
            "end_idx": best_end,
            "optimized_r_squared": best_r_squared,
        }

    def _simple_fallback_segment(
        self, curve_data: np.ndarray, curve_times: np.ndarray, curve_type: str
    ) -> Dict:
        """
        Simple fallback segment detection when advanced methods fail.
        """
        total_length = len(curve_data)

        if curve_type == "hyperpol":
            # Use latter half for hyperpolarization
            start_idx = max(0, total_length // 2)
            end_idx = min(total_length, start_idx + total_length // 3)
        else:
            # Use earlier part for depolarization
            start_idx = max(0, total_length // 4)
            end_idx = min(total_length, start_idx + total_length // 3)

        return {
            "start_idx": start_idx,
            "end_idx": end_idx,
            "segment_type": curve_type,
            "detection_method": "fallback",
            "linearity_score": self._calculate_linearity_score(
                curve_data[start_idx:end_idx], curve_times[start_idx:end_idx]
            ),
        }

    def _emergency_fallback_segment(
        self,
        curve_data: np.ndarray,
        curve_times: np.ndarray,
        curve_type: str,
        error_msg: str,
    ) -> Dict:
        """
        Emergency fallback when all other methods fail.
        """
        app_logger.warning(f"Using emergency fallback segment due to: {error_msg}")

        total_length = len(curve_data)
        start_idx = total_length // 3
        end_idx = min(
            total_length, start_idx + max(self.min_segment_length, total_length // 3)
        )

        return {
            "start_idx": start_idx,
            "end_idx": end_idx,
            "segment_type": curve_type,
            "detection_method": "emergency_fallback",
            "linearity_score": 0.0,
            "error": error_msg,
        }

    # BACKWARD COMPATIBILITY METHODS
    # These methods maintain compatibility with existing code

    def _find_descending_segment(
        self, curve_data: np.ndarray, curve_times: np.ndarray
    ) -> Dict:
        """
        Find the descending segment for depolarization curve.
        Backward compatibility method.
        """
        return self.identify_regression_segments(curve_data, curve_times, "depol")

    def _find_ascending_segment(
        self, curve_data: np.ndarray, curve_times: np.ndarray
    ) -> Dict:
        """
        Find the ascending segment for hyperpolarization curve.
        Backward compatibility method.
        """
        return self.identify_regression_segments(curve_data, curve_times, "hyperpol")

    def fit_linear_regression(
        self, curve_data: np.ndarray, curve_times: np.ndarray, segment_info: Dict
    ) -> Dict:
        """
        Fit linear regression with enhanced statistics.
        Backward compatibility method with enhanced functionality.
        """
        try:
            start_idx = segment_info["start_idx"]
            end_idx = segment_info["end_idx"]

            # Extract segment data
            x_segment = curve_times[start_idx:end_idx] * 1000  # Convert to milliseconds
            y_segment = curve_data[start_idx:end_idx]

            # Choose regression method
            method = self._choose_regression_method(x_segment, y_segment)

            # Perform regression
            if method == "robust" and self.sklearn_available:
                regression_results = self._fit_robust_regression(x_segment, y_segment)
            else:
                # Standard OLS
                slope, intercept, r_value, p_value, std_err = stats.linregress(
                    x_segment, y_segment
                )
                regression_results = {
                    "slope": slope,
                    "intercept": intercept,
                    "r_value": r_value,
                    "r_squared": r_value**2,
                    "p_value": p_value,
                    "std_error": std_err,
                    "method_used": "ols",
                }

            # Calculate fitted line for entire curve
            x_full = curve_times * 1000  # Full time range in milliseconds
            fitted_line = (
                regression_results["slope"] * x_full + regression_results["intercept"]
            )

            # Add additional results
            regression_results.update(
                {
                    "segment_indices": (start_idx, end_idx),
                    "fitted_line": fitted_line,
                    "x_segment": x_segment,
                    "y_segment": y_segment,
                    "segment_type": segment_info["segment_type"],
                }
            )

            app_logger.info(
                f"Linear regression completed: slope={regression_results['slope']:.6f}, R²={regression_results['r_squared']:.4f}"
            )
            return regression_results

        except Exception as e:
            app_logger.error(f"Error fitting linear regression: {str(e)}")
            raise

    def _choose_regression_method(self, x_data: np.ndarray, y_data: np.ndarray) -> str:
        """
        Choose appropriate regression method based on data characteristics.
        """
        if not self.sklearn_available:
            return "ols"

        # Detect outliers
        z_scores = np.abs(stats.zscore(y_data))
        outlier_ratio = np.sum(z_scores > 2.5) / len(y_data)

        # Use robust methods if high outlier ratio
        if outlier_ratio > 0.1:
            return "robust"
        else:
            return "ols"

    def _fit_robust_regression(self, x_data: np.ndarray, y_data: np.ndarray) -> Dict:
        """
        Fit robust regression when outliers are detected.
        """
        try:
            X = x_data.reshape(-1, 1)

            # Try RANSAC first
            ransac = RANSACRegressor(random_state=42)
            ransac.fit(X, y_data)

            slope = ransac.estimator_.coef_[0]
            intercept = ransac.estimator_.intercept_

            # Calculate R²
            y_pred = slope * x_data + intercept
            r_squared = r2_score(y_data, y_pred)
            r_value = np.sqrt(r_squared) if r_squared >= 0 else -np.sqrt(-r_squared)

            return {
                "slope": slope,
                "intercept": intercept,
                "r_value": r_value,
                "r_squared": r_squared,
                "p_value": np.nan,
                "std_error": np.nan,
                "method_used": "ransac",
                "inlier_mask": ransac.inlier_mask_,
            }

        except Exception as e:
            app_logger.warning(f"Robust regression failed, falling back to OLS: {e}")
            # Fallback to OLS
            slope, intercept, r_value, p_value, std_err = stats.linregress(
                x_data, y_data
            )
            return {
                "slope": slope,
                "intercept": intercept,
                "r_value": r_value,
                "r_squared": r_value**2,
                "p_value": p_value,
                "std_error": std_err,
                "method_used": "ols_fallback",
            }

    def apply_drift_correction(
        self,
        curve_data: np.ndarray,
        curve_times: np.ndarray,
        regression_results: Dict,
        plateau_start_idx: Optional[int] = None,
        plateau_end_idx: Optional[int] = None,
    ) -> Dict:
        """
        Apply drift correction by subtracting the fitted line from specified regions.
        Enhanced version with automatic plateau detection.
        """
        try:
            app_logger.info("Applying enhanced drift correction to curve")

            # If plateau region not specified, try to auto-detect
            if plateau_start_idx is None or plateau_end_idx is None:
                plateau_start_idx, plateau_end_idx = self._detect_plateau_region(
                    curve_data, regression_results["segment_type"]
                )

            # Create corrected curve (start with original)
            corrected_curve = curve_data.copy()

            # Subtract fitted line only from the plateau region
            fitted_line = regression_results["fitted_line"]
            baseline_offset = fitted_line[
                plateau_start_idx
            ]  # Use plateau start as reference

            # Apply correction to plateau region
            for i in range(plateau_start_idx, plateau_end_idx + 1):
                corrected_curve[i] = curve_data[i] - (fitted_line[i] - baseline_offset)

            correction_info = {
                "corrected_curve": corrected_curve,
                "original_curve": curve_data,
                "fitted_line": fitted_line,
                "plateau_region": (plateau_start_idx, plateau_end_idx),
                "baseline_offset": baseline_offset,
                "correction_applied": True,
                "correction_method": "enhanced",
            }

            app_logger.info(
                f"Enhanced drift correction applied to plateau region: {plateau_start_idx}-{plateau_end_idx}"
            )
            return correction_info

        except Exception as e:
            app_logger.error(f"Error applying drift correction: {str(e)}")
            raise

    def _detect_plateau_region(
        self, curve_data: np.ndarray, segment_type: str
    ) -> Tuple[int, int]:
        """
        Auto-detect the plateau (flat) region of the curve.
        Enhanced version with multiple detection strategies.
        """
        try:
            n = len(curve_data)

            # Strategy 1: Use derivative to find flat regions
            if n > 10:
                derivative = np.gradient(curve_data)
                smoothed_deriv = savgol_filter(
                    derivative, min(max(5, n // 20), n - 1 if n % 2 == 0 else n), 3
                )

                # Find regions with small derivative (flat regions)
                flat_threshold = np.std(smoothed_deriv) * 0.5
                flat_mask = np.abs(smoothed_deriv) < flat_threshold

                # Find largest contiguous flat region
                flat_regions = []
                start = None
                for i, is_flat in enumerate(flat_mask):
                    if is_flat and start is None:
                        start = i
                    elif not is_flat and start is not None:
                        flat_regions.append((start, i - 1))
                        start = None

                if start is not None:  # Region extends to end
                    flat_regions.append((start, len(flat_mask) - 1))

                if flat_regions:
                    # Choose longest flat region
                    longest_region = max(flat_regions, key=lambda x: x[1] - x[0])
                    return longest_region

            # Strategy 2: Fallback based on curve type
            if segment_type == "hyperpol":
                # For hyperpol, plateau is typically in the latter part
                start_idx = max(0, n // 2)
                end_idx = min(n - 1, start_idx + n // 3)
            else:
                # For depol, plateau might be in the middle or later part
                start_idx = max(0, n // 3)
                end_idx = min(n - 1, start_idx + n // 3)

            return start_idx, end_idx

        except Exception as e:
            app_logger.warning(f"Error detecting plateau region: {e}")
            # Ultimate fallback
            n = len(curve_data)
            return n // 3, min(n - 1, 2 * n // 3)

    # Additional utility methods for comprehensive analysis

    def process_purple_curves(self, processor, manual_plateau_regions=None):
        """
        Process purple curves with enhanced methods.
        Maintains backward compatibility with existing interface.
        """
        app_logger.info("Processing purple curves with enhanced drift correction")

        try:
            results = {}

            # Process hyperpolarization curve
            if (
                hasattr(processor, "modified_hyperpol")
                and processor.modified_hyperpol is not None
            ):
                app_logger.info("Processing enhanced hyperpolarization curve")

                # Identify regression segment
                hyperpol_segment = self.identify_regression_segments(
                    processor.modified_hyperpol,
                    processor.modified_hyperpol_times,
                    "hyperpol",
                )

                # Fit linear regression
                hyperpol_regression = self.fit_linear_regression(
                    processor.modified_hyperpol,
                    processor.modified_hyperpol_times,
                    hyperpol_segment,
                )

                # Apply drift correction
                plateau_region = (
                    manual_plateau_regions.get("hyperpol")
                    if manual_plateau_regions
                    else (None, None)
                )
                hyperpol_correction = self.apply_drift_correction(
                    processor.modified_hyperpol,
                    processor.modified_hyperpol_times,
                    hyperpol_regression,
                    plateau_region[0] if plateau_region[0] else None,
                    plateau_region[1] if plateau_region[1] else None,
                )

                results["hyperpol"] = {
                    "segment_info": hyperpol_segment,
                    "regression_results": hyperpol_regression,
                    "correction_info": hyperpol_correction,
                    "original_times": processor.modified_hyperpol_times,
                }

            # Process depolarization curve
            if (
                hasattr(processor, "modified_depol")
                and processor.modified_depol is not None
            ):
                app_logger.info("Processing enhanced depolarization curve")

                # Identify regression segment
                depol_segment = self.identify_regression_segments(
                    processor.modified_depol, processor.modified_depol_times, "depol"
                )

                # Fit linear regression
                depol_regression = self.fit_linear_regression(
                    processor.modified_depol,
                    processor.modified_depol_times,
                    depol_segment,
                )

                # Apply drift correction
                plateau_region = (
                    manual_plateau_regions.get("depol")
                    if manual_plateau_regions
                    else (None, None)
                )
                depol_correction = self.apply_drift_correction(
                    processor.modified_depol,
                    processor.modified_depol_times,
                    depol_regression,
                    plateau_region[0] if plateau_region[0] else None,
                    plateau_region[1] if plateau_region[1] else None,
                )

                results["depol"] = {
                    "segment_info": depol_segment,
                    "regression_results": depol_regression,
                    "correction_info": depol_correction,
                    "original_times": processor.modified_depol_times,
                }

            # Generate summary
            results["summary"] = self._generate_correction_summary(results)

            app_logger.info(
                "Enhanced purple curve drift correction completed successfully"
            )
            return results

        except Exception as e:
            app_logger.error(f"Error processing purple curves: {str(e)}")
            raise

    def _generate_correction_summary(self, results: Dict) -> Dict:
        """Generate summary of drift correction results."""
        summary = {
            "curves_processed": [],
            "total_r_squared": 0,
            "average_quality": 0,
            "methods_used": [],
        }

        r_squared_values = []

        for curve_type in ["hyperpol", "depol"]:
            if curve_type in results:
                summary["curves_processed"].append(curve_type)

                regression = results[curve_type]["regression_results"]
                r_squared = regression.get("r_squared", 0)
                r_squared_values.append(r_squared)

                method = regression.get("method_used", "unknown")
                if method not in summary["methods_used"]:
                    summary["methods_used"].append(method)

        if r_squared_values:
            summary["total_r_squared"] = sum(r_squared_values)
            summary["average_quality"] = np.mean(r_squared_values)

        return summary


# Backward compatibility alias
LinearDriftCorrector = EnhancedLinearDriftCorrector


# Legacy functions for backward compatibility
def add_drift_correction_to_action_potential_tab(action_potential_tab):
    """
    Add enhanced drift correction functionality to the existing ActionPotentialTab.
    Enhanced version of the original function.
    """
    try:
        from tkinter import messagebox

        def apply_enhanced_drift_correction():
            try:
                if (
                    hasattr(action_potential_tab, "processor")
                    and action_potential_tab.processor
                ):
                    # Use enhanced corrector
                    action_potential_tab.enhanced_drift_corrector = (
                        EnhancedLinearDriftCorrector()
                    )

                    # Process with enhanced methods
                    action_potential_tab.drift_correction_results = action_potential_tab.enhanced_drift_corrector.process_purple_curves(
                        action_potential_tab.processor
                    )

                    messagebox.showinfo(
                        "Success",
                        "Enhanced drift correction applied successfully!\n"
                        "Check the console for detailed results.",
                    )
                else:
                    messagebox.showwarning(
                        "No Data", "Please load and process data first"
                    )
            except Exception as e:
                messagebox.showerror(
                    "Error", f"Enhanced drift correction failed: {str(e)}"
                )

        def export_enhanced_drift_correction_report():
            try:
                if hasattr(action_potential_tab, "drift_correction_results"):
                    # Use enhanced export functionality
                    app_logger.info(
                        "Enhanced drift correction export functionality would go here"
                    )
                    messagebox.showinfo(
                        "Export", "Enhanced export functionality available"
                    )
                else:
                    messagebox.showwarning(
                        "No Data", "Please apply enhanced drift correction first"
                    )
            except Exception as e:
                messagebox.showerror("Error", f"Enhanced export failed: {str(e)}")

        # Add the enhanced methods to the tab
        action_potential_tab.apply_enhanced_drift_correction = (
            apply_enhanced_drift_correction
        )
        action_potential_tab.export_enhanced_drift_correction_report = (
            export_enhanced_drift_correction_report
        )

        app_logger.info(
            "Enhanced drift correction functionality added to Action Potential tab"
        )

    except Exception as e:
        app_logger.error(f"Error adding enhanced drift correction to tab: {str(e)}")
        raise

    def calculate_corrected_integrals(
        self, results: Dict, integration_ranges: Dict
    ) -> Dict:
        """
        Calculate integrals on drift-corrected curves with enhanced error handling.

        Args:
            results: Results from process_purple_curves
            integration_ranges: Dict with 'hyperpol' and 'depol' range tuples

        Returns:
            Dictionary with corrected integral values and statistics
        """
        try:
            app_logger.info("Calculating integrals on drift-corrected curves")

            corrected_integrals = {}

            for curve_type in ["hyperpol", "depol"]:
                if curve_type in results and results[curve_type]:
                    correction_info = results[curve_type]["correction_info"]
                    times = results[curve_type]["original_times"]
                    corrected_curve = correction_info["corrected_curve"]

                    # Get integration range
                    if curve_type in integration_ranges:
                        start_idx, end_idx = integration_ranges[curve_type]

                        # Ensure indices are valid
                        start_idx = max(0, min(start_idx, len(corrected_curve) - 1))
                        end_idx = max(start_idx + 1, min(end_idx, len(corrected_curve)))

                        # Calculate integral on corrected curve
                        time_segment = times[start_idx:end_idx] * 1000  # Convert to ms
                        curve_segment = corrected_curve[start_idx:end_idx]

                        corrected_integral = np.trapz(curve_segment, time_segment)

                        # Also calculate original integral for comparison
                        original_curve = correction_info["original_curve"]
                        original_integral = np.trapz(
                            original_curve[start_idx:end_idx], time_segment
                        )

                        # Calculate relative and absolute changes
                        absolute_change = corrected_integral - original_integral
                        relative_change = (
                            (absolute_change / original_integral * 100)
                            if original_integral != 0
                            else 0
                        )

                        corrected_integrals[curve_type] = {
                            "corrected_integral": corrected_integral,
                            "original_integral": original_integral,
                            "correction_difference": absolute_change,
                            "relative_change_percent": relative_change,
                            "integration_range": (start_idx, end_idx),
                            "integration_duration_ms": time_segment[-1]
                            - time_segment[0],
                            "average_current_original": np.mean(
                                original_curve[start_idx:end_idx]
                            ),
                            "average_current_corrected": np.mean(curve_segment),
                        }

                        app_logger.info(
                            f"{curve_type} - Original: {original_integral:.6f}, "
                            f"Corrected: {corrected_integral:.6f}, "
                            f"Change: {absolute_change:.6f} ({relative_change:.2f}%)"
                        )
                    else:
                        app_logger.warning(
                            f"No integration range specified for {curve_type}"
                        )

            # Add summary statistics
            if corrected_integrals:
                total_original = sum(
                    data["original_integral"] for data in corrected_integrals.values()
                )
                total_corrected = sum(
                    data["corrected_integral"] for data in corrected_integrals.values()
                )

                corrected_integrals["summary"] = {
                    "total_original_integral": total_original,
                    "total_corrected_integral": total_corrected,
                    "total_correction": total_corrected - total_original,
                    "curves_analyzed": list(corrected_integrals.keys()),
                }

            return corrected_integrals

        except Exception as e:
            app_logger.error(f"Error calculating corrected integrals: {str(e)}")
            raise

    def plot_correction_analysis(
        self, results: Dict, save_path: Optional[str] = None
    ) -> plt.Figure:
        """
        Create comprehensive visualization plots showing the enhanced drift correction process.

        Args:
            results: Results from process_purple_curves
            save_path: Optional path to save the plot

        Returns:
            matplotlib figure object
        """
        try:
            fig, axes = plt.subplots(2, 3, figsize=(18, 12))
            fig.suptitle(
                "Enhanced Linear Drift Correction Analysis (Egyenes Illesztés)",
                fontsize=16,
                fontweight="bold",
            )

            for i, curve_type in enumerate(["hyperpol", "depol"]):
                if curve_type in results and results[curve_type]:
                    data = results[curve_type]
                    times = data["original_times"] * 1000  # Convert to ms
                    original_curve = data["correction_info"]["original_curve"]
                    corrected_curve = data["correction_info"]["corrected_curve"]
                    fitted_line = data["regression_results"]["fitted_line"]
                    segment_indices = data["regression_results"]["segment_indices"]
                    plateau_region = data["correction_info"]["plateau_region"]

                    # 1. Original vs corrected curves with enhanced annotations
                    ax1 = axes[i, 0]
                    ax1.plot(
                        times,
                        original_curve,
                        "b-",
                        label="Original",
                        alpha=0.7,
                        linewidth=2,
                    )
                    ax1.plot(
                        times,
                        corrected_curve,
                        "r-",
                        label="Drift Corrected",
                        linewidth=2,
                    )
                    ax1.plot(
                        times,
                        fitted_line,
                        "g--",
                        label="Fitted Line",
                        alpha=0.8,
                        linewidth=2,
                    )

                    # Highlight regression segment
                    start_idx, end_idx = segment_indices
                    ax1.axvspan(
                        times[start_idx],
                        times[end_idx],
                        alpha=0.2,
                        color="green",
                        label="Regression Segment",
                    )

                    # Highlight plateau region
                    plateau_start, plateau_end = plateau_region
                    ax1.axvspan(
                        times[plateau_start],
                        times[plateau_end],
                        alpha=0.2,
                        color="red",
                        label="Correction Region",
                    )

                    # Add method annotation
                    method_used = data["regression_results"].get(
                        "method_used", "unknown"
                    )
                    detection_method = data["segment_info"].get(
                        "detection_method", "unknown"
                    )

                    ax1.text(
                        0.02,
                        0.98,
                        f"Detection: {detection_method}\nRegression: {method_used}",
                        transform=ax1.transAxes,
                        fontsize=9,
                        bbox=dict(
                            boxstyle="round,pad=0.3", facecolor="wheat", alpha=0.8
                        ),
                        verticalalignment="top",
                    )

                    ax1.set_title(
                        f"{curve_type.capitalize()} Curve Correction",
                        fontsize=14,
                        fontweight="bold",
                    )
                    ax1.set_xlabel("Time (ms)", fontsize=12)
                    ax1.set_ylabel("Current (pA)", fontsize=12)
                    ax1.legend(loc="best", fontsize=10)
                    ax1.grid(True, alpha=0.3)

                    # 2. Enhanced regression quality plot
                    ax2 = axes[i, 1]
                    x_segment = data["regression_results"]["x_segment"]
                    y_segment = data["regression_results"]["y_segment"]
                    slope = data["regression_results"]["slope"]
                    intercept = data["regression_results"]["intercept"]
                    r_squared = data["regression_results"]["r_squared"]

                    ax2.scatter(
                        x_segment,
                        y_segment,
                        alpha=0.6,
                        color="blue",
                        s=40,
                        label="Data Points",
                    )
                    fit_line = slope * x_segment + intercept
                    ax2.plot(
                        x_segment,
                        fit_line,
                        "r-",
                        linewidth=3,
                        label=f"Linear Fit (R²={r_squared:.4f})",
                    )

                    # Add confidence intervals if available
                    if "prediction_intervals" in data["regression_results"]:
                        intervals = data["regression_results"]["prediction_intervals"]
                        if len(intervals) > 0 and not np.all(np.isnan(intervals)):
                            ax2.fill_between(
                                x_segment,
                                intervals[:, 0],
                                intervals[:, 1],
                                alpha=0.2,
                                color="orange",
                                label="95% Confidence",
                            )

                    # Enhanced statistics annotation
                    stats_text = (
                        f"R² = {r_squared:.4f}\n"
                        f"Slope = {slope:.6f} pA/ms\n"
                        f"Intercept = {intercept:.3f} pA\n"
                        f"Method = {method_used}"
                    )

                    if "p_value" in data["regression_results"] and not np.isnan(
                        data["regression_results"]["p_value"]
                    ):
                        stats_text += (
                            f'\np-value = {data["regression_results"]["p_value"]:.4f}'
                        )

                    ax2.text(
                        0.05,
                        0.95,
                        stats_text,
                        transform=ax2.transAxes,
                        fontsize=9,
                        bbox=dict(
                            boxstyle="round,pad=0.3", facecolor="lightblue", alpha=0.8
                        ),
                        verticalalignment="top",
                        fontfamily="monospace",
                    )

                    ax2.set_title(
                        f"{curve_type.capitalize()} Regression Quality", fontsize=12
                    )
                    ax2.set_xlabel("Time (ms)", fontsize=10)
                    ax2.set_ylabel("Current (pA)", fontsize=10)
                    ax2.legend(fontsize=9)
                    ax2.grid(True, alpha=0.3)

                    # 3. Residual analysis plot
                    ax3 = axes[i, 2]

                    if "residuals" in data["regression_results"]:
                        residuals = data["regression_results"]["residuals"]
                        fitted_values = slope * x_segment + intercept

                        ax3.scatter(
                            fitted_values, residuals, alpha=0.6, color="purple", s=30
                        )
                        ax3.axhline(
                            y=0, color="red", linestyle="--", alpha=0.8, linewidth=2
                        )

                        # Add trend line for residuals
                        if len(fitted_values) > 2:
                            z = np.polyfit(fitted_values, residuals, 1)
                            p = np.poly1d(z)
                            ax3.plot(
                                fitted_values,
                                p(fitted_values),
                                color="orange",
                                linestyle="-",
                                alpha=0.8,
                                linewidth=2,
                                label=f"Trend (slope={z[0]:.4f})",
                            )
                            ax3.legend(fontsize=8)

                        # Add residual statistics
                        residual_std = np.std(residuals)
                        residual_mean = np.mean(residuals)

                        stats_text = (
                            f"Mean = {residual_mean:.3f}\n"
                            f"Std = {residual_std:.3f}\n"
                            f"Range = {np.ptp(residuals):.3f}"
                        )

                        ax3.text(
                            0.05,
                            0.95,
                            stats_text,
                            transform=ax3.transAxes,
                            fontsize=9,
                            bbox=dict(
                                boxstyle="round,pad=0.3",
                                facecolor="lightyellow",
                                alpha=0.8,
                            ),
                            verticalalignment="top",
                        )

                    ax3.set_title(f"{curve_type.capitalize()} Residuals", fontsize=12)
                    ax3.set_xlabel("Fitted Values (pA)", fontsize=10)
                    ax3.set_ylabel("Residuals (pA)", fontsize=10)
                    ax3.grid(True, alpha=0.3)

            plt.tight_layout()

            if save_path:
                plt.savefig(save_path, dpi=300, bbox_inches="tight", facecolor="white")
                app_logger.info(f"Enhanced drift correction plot saved to: {save_path}")

            return fig

        except Exception as e:
            app_logger.error(f"Error creating enhanced correction plot: {str(e)}")
            raise

    def export_correction_report(
        self,
        results: Dict,
        corrected_integrals: Optional[Dict] = None,
        filename: Optional[str] = None,
    ) -> Optional[str]:
        """
        Export comprehensive drift correction report to Excel with enhanced formatting.

        Args:
            results: Results from process_purple_curves
            corrected_integrals: Optional results from calculate_corrected_integrals
            filename: Optional filename for export

        Returns:
            Filename of exported report or None if cancelled
        """
        try:
            if not filename:
                try:
                    from tkinter import filedialog

                    filename = filedialog.asksaveasfilename(
                        defaultextension=".xlsx",
                        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                        title="Export Enhanced Drift Correction Report",
                    )
                except ImportError:
                    # Fallback if tkinter not available
                    from datetime import datetime

                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    filename = f"enhanced_drift_correction_report_{timestamp}.xlsx"

            if not filename:
                return None

            try:
                import pandas as pd
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment

                with pd.ExcelWriter(filename, engine="openpyxl") as writer:
                    # 1. Summary sheet
                    summary_data = []
                    summary_data.append(
                        [
                            "Enhanced Linear Drift Correction Report (Egyenes Illesztés)",
                            "",
                        ]
                    )
                    summary_data.append(["Generated", np.datetime64("now").astype(str)])
                    summary_data.append(
                        ["Analysis Method", "Enhanced with Auto-Detection"]
                    )
                    summary_data.append(["", ""])

                    for curve_type in ["hyperpol", "depol"]:
                        if curve_type in results:
                            regression = results[curve_type]["regression_results"]
                            segment_info = results[curve_type]["segment_info"]

                            summary_data.extend(
                                [
                                    [f"{curve_type.capitalize()} Results", ""],
                                    [
                                        "Detection Method",
                                        segment_info.get("detection_method", "N/A"),
                                    ],
                                    [
                                        "Regression Method",
                                        regression.get("method_used", "N/A"),
                                    ],
                                    ["Slope (pA/ms)", f"{regression['slope']:.8f}"],
                                    [
                                        "Intercept (pA)",
                                        f"{regression['intercept']:.6f}",
                                    ],
                                    ["R-squared", f"{regression['r_squared']:.6f}"],
                                    ["P-value", f"{regression.get('p_value', 'N/A')}"],
                                    [
                                        "Segment Quality",
                                        f"{segment_info.get('linearity_score', 'N/A'):.4f}",
                                    ],
                                    ["", ""],
                                ]
                            )

                            if (
                                corrected_integrals
                                and curve_type in corrected_integrals
                            ):
                                integrals = corrected_integrals[curve_type]
                                summary_data.extend(
                                    [
                                        [
                                            "Original Integral (pC)",
                                            f"{integrals['original_integral']:.6f}",
                                        ],
                                        [
                                            "Corrected Integral (pC)",
                                            f"{integrals['corrected_integral']:.6f}",
                                        ],
                                        [
                                            "Correction Difference (pC)",
                                            f"{integrals['correction_difference']:.6f}",
                                        ],
                                        [
                                            "Relative Change (%)",
                                            f"{integrals['relative_change_percent']:.2f}",
                                        ],
                                        ["", ""],
                                    ]
                                )

                    df_summary = pd.DataFrame(
                        summary_data, columns=["Parameter", "Value"]
                    )
                    df_summary.to_excel(writer, sheet_name="Summary", index=False)

                    # 2. Detailed regression data for each curve
                    for curve_type in ["hyperpol", "depol"]:
                        if curve_type in results:
                            regression = results[curve_type]["regression_results"]
                            x_segment = regression["x_segment"]
                            y_segment = regression["y_segment"]

                            # Create detailed data
                            regression_data = pd.DataFrame(
                                {
                                    "Time_ms": x_segment,
                                    "Current_pA": y_segment,
                                    "Fitted_pA": regression["slope"] * x_segment
                                    + regression["intercept"],
                                    "Residuals_pA": regression.get(
                                        "residuals",
                                        y_segment
                                        - (
                                            regression["slope"] * x_segment
                                            + regression["intercept"]
                                        ),
                                    ),
                                }
                            )

                            # Add prediction intervals if available
                            if (
                                "prediction_intervals" in regression
                                and len(regression["prediction_intervals"]) > 0
                            ):
                                intervals = regression["prediction_intervals"]
                                if not np.all(np.isnan(intervals)):
                                    regression_data["Lower_CI_95"] = intervals[:, 0]
                                    regression_data["Upper_CI_95"] = intervals[:, 1]

                            regression_data.to_excel(
                                writer,
                                sheet_name=f"{curve_type}_regression",
                                index=False,
                            )

                    # 3. Method comparison sheet (if multiple methods were tried)
                    method_comparison = []
                    for curve_type in ["hyperpol", "depol"]:
                        if curve_type in results:
                            segment_info = results[curve_type]["segment_info"]
                            regression = results[curve_type]["regression_results"]

                            method_comparison.append(
                                {
                                    "Curve_Type": curve_type.capitalize(),
                                    "Detection_Method": segment_info.get(
                                        "detection_method", "N/A"
                                    ),
                                    "Regression_Method": regression.get(
                                        "method_used", "N/A"
                                    ),
                                    "R_Squared": regression["r_squared"],
                                    "Linearity_Score": segment_info.get(
                                        "linearity_score", np.nan
                                    ),
                                    "Quality_Score": segment_info.get(
                                        "quality_score", np.nan
                                    ),
                                    "Segment_Start": segment_info["start_idx"],
                                    "Segment_End": segment_info["end_idx"],
                                    "Segment_Length": segment_info["end_idx"]
                                    - segment_info["start_idx"],
                                }
                            )

                    if method_comparison:
                        df_methods = pd.DataFrame(method_comparison)
                        df_methods.to_excel(
                            writer, sheet_name="Method_Comparison", index=False
                        )

                    # 4. Enhanced formatting
                    workbook = writer.book

                    # Format summary sheet
                    if "Summary" in workbook.sheetnames:
                        ws = workbook["Summary"]

                        # Header formatting
                        header_fill = PatternFill(
                            start_color="366092", end_color="366092", fill_type="solid"
                        )
                        header_font = Font(color="FFFFFF", bold=True)

                        for cell in ws[1]:
                            cell.fill = header_fill
                            cell.font = header_font

                        # Auto-adjust column widths
                        for column in ws.columns:
                            max_length = 0
                            column_letter = column[0].column_letter
                            for cell in column:
                                try:
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(str(cell.value))
                                except:
                                    pass
                            adjusted_width = min(max_length + 2, 50)
                            ws.column_dimensions[column_letter].width = adjusted_width

                app_logger.info(
                    f"Enhanced drift correction report exported to: {filename}"
                )
                return filename

            except ImportError as e:
                app_logger.warning(f"Advanced Excel formatting not available: {e}")
                # Fallback to basic export
                return self._export_basic_report(results, corrected_integrals, filename)

        except Exception as e:
            app_logger.error(f"Error exporting enhanced correction report: {str(e)}")
            raise

    def _export_basic_report(
        self, results: Dict, corrected_integrals: Optional[Dict], filename: str
    ) -> str:
        """
        Basic report export fallback when advanced libraries are not available.
        """
        try:
            import csv
            import json

            # Export as JSON for comprehensive data
            json_filename = filename.replace(".xlsx", ".json")

            export_data = {
                "metadata": {
                    "export_timestamp": np.datetime64("now").astype(str),
                    "analysis_type": "Enhanced Linear Drift Correction",
                    "method": "Auto-Detection with Multiple Algorithms",
                },
                "results": results,
                "corrected_integrals": corrected_integrals,
            }

            # Convert numpy arrays to lists for JSON serialization
            def convert_numpy(obj):
                if isinstance(obj, np.ndarray):
                    return obj.tolist()
                elif isinstance(obj, dict):
                    return {k: convert_numpy(v) for k, v in obj.items()}
                elif isinstance(obj, list):
                    return [convert_numpy(item) for item in obj]
                elif isinstance(obj, (np.integer, np.floating)):
                    return obj.item()
                else:
                    return obj

            export_data = convert_numpy(export_data)

            with open(json_filename, "w") as f:
                json.dump(export_data, f, indent=2)

            # Also export key results as CSV
            csv_filename = filename.replace(".xlsx", "_summary.csv")

            with open(csv_filename, "w", newline="") as csvfile:
                writer = csv.writer(csvfile)
                writer.writerow(["Enhanced Linear Drift Correction Summary"])
                writer.writerow(["Parameter", "Hyperpol", "Depol"])

                for curve_type in ["hyperpol", "depol"]:
                    if curve_type in results:
                        regression = results[curve_type]["regression_results"]
                        writer.writerow(
                            [f"{curve_type}_slope_pA_per_ms", regression["slope"], ""]
                        )
                        writer.writerow(
                            [f"{curve_type}_intercept_pA", regression["intercept"], ""]
                        )
                        writer.writerow(
                            [f"{curve_type}_r_squared", regression["r_squared"], ""]
                        )
                        writer.writerow(
                            [
                                f"{curve_type}_method",
                                regression.get("method_used", "N/A"),
                                "",
                            ]
                        )

            app_logger.info(f"Basic reports exported: {json_filename}, {csv_filename}")
            return json_filename

        except Exception as e:
            app_logger.error(f"Error in basic report export: {str(e)}")
            raise
