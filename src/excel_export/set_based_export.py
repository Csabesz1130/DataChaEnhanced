"""
Set-Based Excel Export for DataChaEnhanced
==========================================
Location: src/excel_export/set_based_export.py

This module provides batch export functionality for multiple files organized by set number,
creating separate Excel files for each set with individual sheets per file.
"""

import os
import re
from datetime import datetime
from typing import Dict, List, Any, Optional
from tkinter import filedialog, messagebox
from openpyxl.styles import Font, PatternFill
from src.utils.logger import app_logger
from .curve_analysis_export import export_curve_analysis_to_excel


def export_sets_to_excel(app) -> bool:
    """Export multiple files organized by set number to Excel files.
    
    Args:
        app: Main application instance with action_potential_processor
        
    Returns:
        bool: True if export successful, False otherwise
    """
    try:
        # Select folder containing ATF files
        folder_path = filedialog.askdirectory(
            title="Select folder containing ATF files for set-based export"
        )
        
        if not folder_path:
            return False  # User cancelled
        
        # Find all ATF files in the folder
        atf_files = _find_atf_files(folder_path)
        if not atf_files:
            messagebox.showwarning("No Files", "No ATF files found in the selected folder.")
            return False
        
        # Organize files by set number
        sets_data = _organize_files_by_set(atf_files)
        if not sets_data:
            messagebox.showwarning("No Sets", "No valid sets found in the files.")
            return False
        
        # Select output directory
        output_dir = filedialog.askdirectory(
            title="Select output directory for Excel files"
        )
        
        if not output_dir:
            return False  # User cancelled
        
        # Create ExportedSets subdirectory
        export_dir = os.path.join(output_dir, "ExportedSets")
        os.makedirs(export_dir, exist_ok=True)
        
        # Export each set
        exported_files = []
        for set_number, files in sets_data.items():
            success = _export_single_set(app, set_number, files, export_dir)
            if success:
                exported_files.append(f"Set_{set_number}")
        
        if exported_files:
            messagebox.showinfo(
                "Export Complete", 
                f"Successfully exported {len(exported_files)} sets:\n" + 
                "\n".join(exported_files)
            )
            app_logger.info(f"Set-based export completed: {len(exported_files)} sets")
            return True
        else:
            messagebox.showerror("Export Failed", "No sets were successfully exported.")
            return False
            
    except Exception as e:
        app_logger.error(f"Set-based export failed: {str(e)}")
        messagebox.showerror("Export Error", f"An error occurred during set-based export:\n{str(e)}")
        return False


def _find_atf_files(folder_path: str) -> List[str]:
    """Find all ATF files in the specified folder.
    
    Args:
        folder_path: Path to the folder to search
        
    Returns:
        List[str]: List of ATF file paths
    """
    atf_files = []
    
    try:
        for filename in os.listdir(folder_path):
            if filename.lower().endswith('.atf'):
                file_path = os.path.join(folder_path, filename)
                atf_files.append(file_path)
        
        app_logger.info(f"Found {len(atf_files)} ATF files in {folder_path}")
        return atf_files
        
    except Exception as e:
        app_logger.error(f"Failed to find ATF files: {str(e)}")
        return []


def _organize_files_by_set(atf_files: List[str]) -> Dict[int, List[Dict[str, Any]]]:
    """Organize files by set number based on filename pattern.
    
    Args:
        atf_files: List of ATF file paths
        
    Returns:
        Dict[int, List[Dict]]: Dictionary mapping set numbers to file data
    """
    sets_data = {}
    
    for file_path in atf_files:
        filename = os.path.basename(file_path)
        
        # Parse filename pattern: 0057_0_-100.atf
        # Extract: file_number, set_number, voltage
        match = re.match(r'(\d+)_(\d+)_(-?\d+)\.atf', filename)
        
        if match:
            file_number = match.group(1)
            set_number = int(match.group(2))
            voltage = int(match.group(3))
            
            file_data = {
                'file_path': file_path,
                'filename': filename,
                'file_number': file_number,
                'set_number': set_number,
                'voltage': voltage
            }
            
            if set_number not in sets_data:
                sets_data[set_number] = []
            
            sets_data[set_number].append(file_data)
            app_logger.debug(f"Organized file {filename} into set {set_number}")
        else:
            app_logger.warning(f"Could not parse filename pattern: {filename}")
    
    # Sort files within each set by file number
    for set_number in sets_data:
        sets_data[set_number].sort(key=lambda x: x['file_number'])
    
    app_logger.info(f"Organized files into {len(sets_data)} sets")
    return sets_data


def _export_single_set(app, set_number: int, files: List[Dict[str, Any]], output_dir: str) -> bool:
    """Export a single set to an Excel file.
    
    Args:
        app: Main application instance
        set_number: Set number for this export
        files: List of file data dictionaries
        output_dir: Output directory for Excel files
        
    Returns:
        bool: True if export successful, False otherwise
    """
    try:
        from datetime import datetime
        
        # Generate timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        
        # Create Excel filename
        excel_filename = f"Set_{set_number}_{timestamp}.xlsx"
        excel_path = os.path.join(output_dir, excel_filename)
        
        # Create workbook
        import openpyxl
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Process each file in the set
        for file_data in files:
            sheet_name = file_data['file_number']
            
            # Create sheet for this file
            ws = wb.create_sheet(sheet_name)
            
            # Add file information
            _add_file_info_to_sheet(ws, file_data)
            
            # Add fitting data (if available)
            _add_fitting_data_to_sheet(ws, file_data, app)
            
            # Add integration data (if available)
            _add_integration_data_to_sheet(ws, file_data, app)
            
            # Add capacitance data (if available)
            _add_capacitance_data_to_sheet(ws, file_data, app)
            
            # Add summary
            _add_summary_to_sheet(ws, file_data, app)
        
        # Save workbook
        wb.save(excel_path)
        app_logger.info(f"Exported set {set_number} to {excel_path}")
        return True
        
    except Exception as e:
        app_logger.error(f"Failed to export set {set_number}: {str(e)}")
        return False


def _add_file_info_to_sheet(ws, file_data: Dict[str, Any]):
    """Add file information to worksheet."""
    # Header styling
    header_font = Font(bold=True, size=14)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Title
    ws['A1'] = "FILE INFORMATION"
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws.merge_cells('A1:B1')
    
    # File information
    info_data = [
        ("File Name", file_data['filename']),
        ("File Number", file_data['file_number']),
        ("Set Number", file_data['set_number']),
        ("Voltage (mV)", file_data['voltage']),
        ("Export Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    ]
    
    for i, (label, value) in enumerate(info_data, start=3):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = str(value)


def _add_fitting_data_to_sheet(ws, file_data: Dict[str, Any], app):
    """Add fitting data to worksheet."""
    # This would need to be implemented based on how fitting data is stored
    # For now, add placeholder
    ws['A10'] = "FITTING DATA"
    ws['A10'].font = Font(bold=True, size=12)
    ws['A11'] = "Note: Fitting data would be populated here"
    ws['A11'].font = Font(italic=True)


def _add_integration_data_to_sheet(ws, file_data: Dict[str, Any], app):
    """Add integration data to worksheet."""
    # This would need to be implemented based on how integration data is stored
    # For now, add placeholder
    ws['A15'] = "INTEGRATION DATA"
    ws['A15'].font = Font(bold=True, size=12)
    ws['A16'] = "Note: Integration data would be populated here"
    ws['A16'].font = Font(italic=True)


def _add_capacitance_data_to_sheet(ws, file_data: Dict[str, Any], app):
    """Add capacitance data to worksheet."""
    # This would need to be implemented based on how capacitance data is stored
    # For now, add placeholder
    ws['A20'] = "CAPACITANCE DATA"
    ws['A20'].font = Font(bold=True, size=12)
    ws['A21'] = "Note: Capacitance data would be populated here"
    ws['A21'].font = Font(italic=True)


def _add_summary_to_sheet(ws, file_data: Dict[str, Any], app):
    """Add summary to worksheet."""
    # This would need to be implemented based on available data
    # For now, add placeholder
    ws['A25'] = "SUMMARY"
    ws['A25'].font = Font(bold=True, size=12)
    ws['A26'] = "Note: Summary data would be populated here"
    ws['A26'].font = Font(italic=True)


def _get_fitting_results_for_file(file_data: Dict[str, Any], app) -> Dict[str, Any]:
    """Get fitting results for a specific file.
    
    This is a placeholder function that would need to be implemented
    based on how fitting data is stored and accessed.
    
    Args:
        file_data: File data dictionary
        app: Main application instance
        
    Returns:
        Dict[str, Any]: Fitting results dictionary
    """
    # This would need to be implemented based on the actual data structure
    # For now, return empty results
    return {
        'hyperpol': {
            'linear': {},
            'exponential': {}
        },
        'depol': {
            'linear': {},
            'exponential': {}
        }
    }


def _get_integration_data_for_file(file_data: Dict[str, Any], app) -> Dict[str, Any]:
    """Get integration data for a specific file.
    
    This is a placeholder function that would need to be implemented
    based on how integration data is stored and accessed.
    
    Args:
        file_data: File data dictionary
        app: Main application instance
        
    Returns:
        Dict[str, Any]: Integration data dictionary
    """
    # This would need to be implemented based on the actual data structure
    # For now, return empty results
    return {
        'hyperpol': {},
        'depol': {}
    }


def _get_capacitance_data_for_file(file_data: Dict[str, Any], app) -> Dict[str, Any]:
    """Get capacitance data for a specific file.
    
    This is a placeholder function that would need to be implemented
    based on how capacitance data is stored and accessed.
    
    Args:
        file_data: File data dictionary
        app: Main application instance
        
    Returns:
        Dict[str, Any]: Capacitance data dictionary
    """
    # This would need to be implemented based on the actual data structure
    # For now, return empty results
    return {
        'linear_capacitance': 'N/A',
        'method': 'Not calculated',
        'notes': 'Capacitance calculation not implemented'
    }
