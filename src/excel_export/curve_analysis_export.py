"""
Curve Analysis Excel Export for DataChaEnhanced
===============================================
Location: src/excel_export/curve_analysis_export.py

This module provides Excel export functionality for curve fitting analysis results,
creating a comprehensive 6-sheet Excel workbook with all fitting parameters.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, Any
from src.utils.logger import app_logger


def export_curve_analysis_to_excel(export_data: Dict[str, Any], filepath: str) -> bool:
    """Export curve analysis data to Excel with 6 sheets.
    
    Args:
        export_data: Dictionary containing all export data
        filepath: Path where to save the Excel file
        
    Returns:
        bool: True if export successful, False otherwise
    """
    try:
        # Create new workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create all 6 sheets
        _create_file_information_sheet(wb, export_data)
        _create_linear_fitting_sheet(wb, export_data)
        _create_exponential_fitting_sheet(wb, export_data)
        _create_integration_sheet(wb, export_data)
        _create_capacitance_sheet(wb, export_data)
        _create_summary_sheet(wb, export_data)
        
        # Save workbook
        wb.save(filepath)
        app_logger.info(f"Excel export completed successfully: {filepath}")
        return True
        
    except Exception as e:
        app_logger.error(f"Failed to export to Excel: {str(e)}")
        return False


def _create_file_information_sheet(wb: openpyxl.Workbook, export_data: Dict[str, Any]):
    """Create File Information sheet."""
    ws = wb.create_sheet("File Information")
    
    # Header styling
    header_font = Font(bold=True, size=14)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Data styling
    data_font = Font(size=11)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    # Title
    ws['A1'] = "FILE INFORMATION"
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = header_alignment
    ws.merge_cells('A1:B1')
    
    # File information
    info_data = [
        ("File Name", export_data.get('filename', 'N/A')),
        ("Voltage (mV)", export_data.get('voltage', 'N/A')),
        ("Export Date", export_data.get('export_date', 'N/A')),
        ("Analysis Type", export_data.get('analysis_type', 'Curve Fitting')),
        ("File Path", export_data.get('file_path', 'N/A'))
    ]
    
    for i, (label, value) in enumerate(info_data, start=3):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = str(value)
        
        # Apply styling
        for cell in [ws[f'A{i}'], ws[f'B{i}']]:
            cell.font = data_font
            cell.border = border
    
    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30


def _create_linear_fitting_sheet(wb: openpyxl.Workbook, export_data: Dict[str, Any]):
    """Create Linear Fitting sheet."""
    ws = wb.create_sheet("Linear Fitting")
    
    # Header styling
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Title
    ws['A1'] = "LINEAR FITTING"
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = header_alignment
    ws.merge_cells('A1:D1')
    
    # Subheaders
    ws['A3'] = "Parameter"
    ws['B3'] = "Hyperpol"
    ws['C3'] = "Depol"
    ws['D3'] = "Notes"
    
    for cell in [ws['A3'], ws['B3'], ws['C3'], ws['D3']]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    # Get fitting data
    fitting_results = export_data.get('fitting_results', {})
    
    # Linear fitting parameters
    linear_data = [
        ("Slope (pA/ms)", "linear", "slope"),
        ("Y-intercept (pA)", "linear", "intercept"),
        ("R²", "linear", "r_squared"),
        ("Equation", "linear", "equation")
    ]
    
    row = 4
    for param_name, fit_type, param_key in linear_data:
        ws[f'A{row}'] = param_name
        
        # Hyperpol data
        hyperpol_data = fitting_results.get('hyperpol', {}).get(fit_type, {})
        ws[f'B{row}'] = _format_fitting_value(hyperpol_data.get(param_key))
        
        # Depol data
        depol_data = fitting_results.get('depol', {}).get(fit_type, {})
        ws[f'C{row}'] = _format_fitting_value(depol_data.get(param_key))
        
        ws[f'D{row}'] = _get_linear_notes(param_name)
        row += 1
    
    # Auto-adjust column widths
    for col in ['A', 'B', 'C', 'D']:
        ws.column_dimensions[col].width = 20


def _create_exponential_fitting_sheet(wb: openpyxl.Workbook, export_data: Dict[str, Any]):
    """Create Exponential Fitting sheet."""
    ws = wb.create_sheet("Exponential Fitting")
    
    # Header styling
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Title
    ws['A1'] = "EXPONENTIAL FITTING"
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = header_alignment
    ws.merge_cells('A1:D1')
    
    # Subheaders
    ws['A3'] = "Parameter"
    ws['B3'] = "Hyperpol"
    ws['C3'] = "Depol"
    ws['D3'] = "Notes"
    
    for cell in [ws['A3'], ws['B3'], ws['C3'], ws['D3']]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    # Get fitting data
    fitting_results = export_data.get('fitting_results', {})
    
    # Exponential fitting parameters
    exp_data = [
        ("Tau (ms)", "exponential", "tau_ms"),
        ("Amplitude (pA)", "exponential", "A"),
        ("Offset (pA)", "exponential", "C"),
        ("R²", "exponential", "r_squared"),
        ("Model Type", "exponential", "model_type")
    ]
    
    row = 4
    for param_name, fit_type, param_key in exp_data:
        ws[f'A{row}'] = param_name
        
        # Hyperpol data
        hyperpol_data = fitting_results.get('hyperpol', {}).get(fit_type, {})
        ws[f'B{row}'] = _format_fitting_value(hyperpol_data.get(param_key))
        
        # Depol data
        depol_data = fitting_results.get('depol', {}).get(fit_type, {})
        ws[f'C{row}'] = _format_fitting_value(depol_data.get(param_key))
        
        ws[f'D{row}'] = _get_exponential_notes(param_name)
        row += 1
    
    # Auto-adjust column widths
    for col in ['A', 'B', 'C', 'D']:
        ws.column_dimensions[col].width = 20


def _create_integration_sheet(wb: openpyxl.Workbook, export_data: Dict[str, Any]):
    """Create Integration sheet."""
    ws = wb.create_sheet("Integration")
    
    # Header styling
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Title
    ws['A1'] = "INTEGRATION"
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = header_alignment
    ws.merge_cells('A1:D1')
    
    # Subheaders
    ws['A3'] = "Parameter"
    ws['B3'] = "Hyperpol"
    ws['C3'] = "Depol"
    ws['D3'] = "Notes"
    
    for cell in [ws['A3'], ws['B3'], ws['C3'], ws['D3']]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    # Integration data
    integration_data = export_data.get('integration_data', {})
    
    integration_params = [
        ("Integral (pA·s)", "integral_value"),
        ("Start Point Index", "start_index"),
        ("End Point Index", "end_index"),
        ("Integration Method", "method")
    ]
    
    row = 4
    for param_name, param_key in integration_params:
        ws[f'A{row}'] = param_name
        
        # Hyperpol data
        hyperpol_integration = integration_data.get('hyperpol', {})
        ws[f'B{row}'] = _format_fitting_value(hyperpol_integration.get(param_key))
        
        # Depol data
        depol_integration = integration_data.get('depol', {})
        ws[f'C{row}'] = _format_fitting_value(depol_integration.get(param_key))
        
        ws[f'D{row}'] = _get_integration_notes(param_name)
        row += 1
    
    # Auto-adjust column widths
    for col in ['A', 'B', 'C', 'D']:
        ws.column_dimensions[col].width = 20


def _create_capacitance_sheet(wb: openpyxl.Workbook, export_data: Dict[str, Any]):
    """Create Capacitance sheet."""
    ws = wb.create_sheet("Capacitance")
    
    # Header styling
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Title
    ws['A1'] = "CAPACITANCE"
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = header_alignment
    ws.merge_cells('A1:B1')
    
    # Capacitance data
    capacitance_data = export_data.get('capacitance_data', {})
    
    capacitance_params = [
        ("Linear Capacitance (nF)", capacitance_data.get('linear_capacitance', 'N/A')),
        ("Calculation Method", capacitance_data.get('method', 'Linear Fit')),
        ("Notes", capacitance_data.get('notes', 'Calculated from linear fitting parameters'))
    ]
    
    row = 3
    for param_name, value in capacitance_params:
        ws[f'A{row}'] = param_name
        ws[f'B{row}'] = str(value)
        row += 1
    
    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 30


def _create_summary_sheet(wb: openpyxl.Workbook, export_data: Dict[str, Any]):
    """Create Summary sheet."""
    ws = wb.create_sheet("Summary")
    
    # Header styling
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Title
    ws['A1'] = "SUMMARY"
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].alignment = header_alignment
    ws.merge_cells('A1:B1')
    
    # Get all data
    fitting_results = export_data.get('fitting_results', {})
    integration_data = export_data.get('integration_data', {})
    capacitance_data = export_data.get('capacitance_data', {})
    
    # Summary sections
    sections = [
        ("=== HYPERPOLARIZATION ===", ""),
        ("Linear Slope (pA/ms)", _get_linear_slope(fitting_results, 'hyperpol')),
        ("Linear Y0 (pA)", _get_linear_intercept(fitting_results, 'hyperpol')),
        ("Linear R²", _get_linear_r2(fitting_results, 'hyperpol')),
        ("Integral (pA·s)", _get_integral_value(integration_data, 'hyperpol')),
        ("", ""),
        ("=== DEPOLARIZATION ===", ""),
        ("Linear Slope (pA/ms)", _get_linear_slope(fitting_results, 'depol')),
        ("Linear Y0 (pA)", _get_linear_intercept(fitting_results, 'depol')),
        ("Linear R²", _get_linear_r2(fitting_results, 'depol')),
        ("Integral (pA·s)", _get_integral_value(integration_data, 'depol')),
        ("", ""),
        ("=== OVERALL ===", ""),
        ("Linear Capacitance (nF)", capacitance_data.get('linear_capacitance', 'N/A'))
    ]
    
    row = 3
    for param_name, value in sections:
        ws[f'A{row}'] = param_name
        ws[f'B{row}'] = str(value)
        
        # Style section headers
        if param_name.startswith("==="):
            ws[f'A{row}'].font = Font(bold=True, size=12)
            ws[f'A{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        
        row += 1
    
    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20


def _format_fitting_value(value) -> str:
    """Format fitting value for display."""
    if value is None:
        return "N/A"
    elif isinstance(value, (int, float)):
        return f"{value:.6f}"
    else:
        return str(value)


def _get_linear_notes(param_name: str) -> str:
    """Get notes for linear fitting parameters."""
    notes = {
        "Slope (pA/ms)": "Rate of change",
        "Y-intercept (pA)": "Baseline value",
        "R²": "Goodness of fit (0-1)",
        "Equation": "Linear equation"
    }
    return notes.get(param_name, "")


def _get_exponential_notes(param_name: str) -> str:
    """Get notes for exponential fitting parameters."""
    notes = {
        "Tau (ms)": "Time constant",
        "Amplitude (pA)": "Peak amplitude",
        "Offset (pA)": "Baseline offset",
        "R²": "Goodness of fit (0-1)",
        "Model Type": "decay or growth"
    }
    return notes.get(param_name, "")


def _get_integration_notes(param_name: str) -> str:
    """Get notes for integration parameters."""
    notes = {
        "Integral (pA·s)": "Area under curve",
        "Start Point Index": "Integration start",
        "End Point Index": "Integration end",
        "Integration Method": "Calculation method"
    }
    return notes.get(param_name, "")


def _get_linear_slope(fitting_results: Dict, curve_type: str) -> str:
    """Get linear slope for summary."""
    linear_data = fitting_results.get(curve_type, {}).get('linear', {})
    return _format_fitting_value(linear_data.get('slope'))


def _get_linear_intercept(fitting_results: Dict, curve_type: str) -> str:
    """Get linear intercept for summary."""
    linear_data = fitting_results.get(curve_type, {}).get('linear', {})
    return _format_fitting_value(linear_data.get('intercept'))


def _get_linear_r2(fitting_results: Dict, curve_type: str) -> str:
    """Get linear R² for summary."""
    linear_data = fitting_results.get(curve_type, {}).get('linear', {})
    return _format_fitting_value(linear_data.get('r_squared'))


def _get_integral_value(integration_data: Dict, curve_type: str) -> str:
    """Get integral value for summary."""
    curve_integration = integration_data.get(curve_type, {})
    return _format_fitting_value(curve_integration.get('integral_value'))
